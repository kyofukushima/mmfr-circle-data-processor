import streamlit as st
import pandas as pd
import io
import time
import os
import chardet
import re
import datetime
import hashlib
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from copy import copy
from openai import OpenAI

# テンプレートファイルのパスを環境変数から取得（テスト時に切り替え可能）
TEMPLATE_FILE = os.getenv('TEMPLATE_FILE', 'template.xlsx')

def normalize_value(raw_value):
    """欠損値を統一的に処理する共通関数
    
    Args:
        raw_value: 処理対象の値
    
    Returns:
        str: 正規化された値（空文字列または文字列）
    """
    if pd.isna(raw_value):
        return ''
    value = str(raw_value).strip()
    if value in ['nan', 'None', '<NA>']:
        return ''
    return value

def get_excel_column_name(column_index):
    """列のインデックスをエクセル形式のアルファベットに変換する
    
    Args:
        column_index (int): 列のインデックス（0始まり）
    
    Returns:
        str: エクセル形式の列名（A, B, C, ..., AA, AB, ...）
    """
    result = ""
    while column_index >= 0:
        result = chr(column_index % 26 + ord('A')) + result
        column_index = column_index // 26 - 1
    return result

def get_column_position_text(df, column_name):
    """データフレームの列名から列位置テキストを生成する
    
    Args:
        df (pd.DataFrame): 対象のデータフレーム
        column_name (str): 列名
    
    Returns:
        str: 列位置テキスト（例: "（BC列）"）
    """
    try:
        column_index = df.columns.get_loc(column_name)
        excel_column = get_excel_column_name(column_index)
        return f"（{excel_column}列）"
    except KeyError:
        return "（不明列）"

def detect_encoding(file_content):
    """ファイルのエンコーディングを検出する"""
    result = chardet.detect(file_content)
    return result['encoding'], result['confidence']

def process_binary_columns(df):
    """0/1の値を持つ列の変換処理を行う
    
    Args:
        df (pd.DataFrame): 処理対象のデータフレーム
    
    Returns:
        tuple: (処理後のデータフレーム, 処理内容のデータフレーム)
    """
    # 変換対象の列を定義
    target_patterns = [
        '対象年齢',  # 対象年齢を含む列
        '要会費',
        '冊子掲載可',
        'HP掲載可',
        'オープンデータ掲載可'
    ]
    
    # 変換対象の列名を抽出
    target_columns = []
    for pattern in target_patterns:
        matched_columns = [col for col in df.columns if pattern in str(col)]
        target_columns.extend(matched_columns)
    
    # 重複を削除
    target_columns = list(dict.fromkeys(target_columns))
    
    # 処理内容をデータフレームとして作成
    process_df = pd.DataFrame(columns=['処理内容', '対象列'])
    if target_columns:
        for col in target_columns:
            # 列を文字列型に変換
            df[col] = df[col].astype(str)
            # 0を空欄に変換
            df[col] = df[col].replace('0', '')
            # 1を○に変換
            df[col] = df[col].replace('1', '○')
        
        # 処理内容のデータフレームを作成
        process_df = pd.DataFrame({
            '処理内容': ['0を空欄に変換および1を○に変換'],
            '対象列': [', '.join(target_columns)]
        })
    
    return df, process_df

def add_location_column(circle_data,df_f):
    """
    場所列の追加施設情報のエクスポートデータを抽出する
    S列「活動場所」の施設名称を参考に、施設情報のエクスポートデータからJ列「場所」を抽出・突合する
    J列「場所」から抽出・突合した情報をAY列の「場所」に入力する
    S列「活動場所」に施設名称がなかったり【育児サークル等地区表示用】○○区を指定している場合にはAY列の「場所」は空欄になるが、この場合空欄になるのが正。
    
    Returns:
        tuple: (処理後のデータフレーム, 処理内容のデータフレーム)
    """
    # 場所情報の追加（施設情報データは既に検証済み）
    circle_data['場所'] = circle_data['活動場所'].map(df_f.set_index('施設名')['場所'])
    
    # 処理内容のデータフレームを作成
    process_df = pd.DataFrame({
        '処理内容': ['活動場所の施設名称から場所情報を抽出・突合'],
        '対象列': ['活動場所 → 場所']
    })
    
    return circle_data, process_df

def check_data_consistency(circle_data, last_month_data):
    """
    育児サークルデータと先月分データの整合性をチェックする
    
    Args:
        circle_data (pd.DataFrame): 育児サークルデータ
        last_month_data (pd.DataFrame): 先月分データ
    
    Returns:
        None
    
    Raises:
        st.stop(): データの不一致がある場合に処理を停止
    """
    error_messages = []
    
    # 必要な列の存在チェック
    required_columns = ['スラッグ', 'サークル名']
    
    for col in required_columns:
        if col not in circle_data.columns:
            error_messages.append(f"### 育児サークルデータに「{col}」列が存在しません")
        if col not in last_month_data.columns:
            error_messages.append(f"### 先月分データに「{col}」列が存在しません")
    
    if error_messages:
        st.error('\n'.join(error_messages))
        st.stop()
    
    # スラッグの重複チェック
    circle_duplicates = circle_data[circle_data['スラッグ'].duplicated()]['スラッグ'].unique()
    last_month_duplicates = last_month_data[last_month_data['スラッグ'].duplicated()]['スラッグ'].unique()
    
    if len(circle_duplicates) > 0:
        error_messages.append("### 育児サークルデータ内で重複しているスラッグ:")
        for slug in circle_duplicates:
            duplicate_rows = circle_data[circle_data['スラッグ'] == slug]
            error_messages.append(f"- スラッグ: {slug}")
            for idx, row in duplicate_rows.iterrows():
                circle_name = row.get('サークル名', '不明')
                error_messages.append(f"  - 行{idx+1}: {circle_name}")
    
    if len(last_month_duplicates) > 0:
        error_messages.append("\n### 先月分データ内で重複しているスラッグ:")
        for slug in last_month_duplicates:
            duplicate_rows = last_month_data[last_month_data['スラッグ'] == slug]
            error_messages.append(f"- スラッグ: {slug}")
            for idx, row in duplicate_rows.iterrows():
                circle_name = row.get('サークル名', '不明')
                error_messages.append(f"  - 行{idx+1}: {circle_name}")
    
    # スラッグの存在チェック（空欄・欠損値を除外）
    circle_slugs = set(circle_data['スラッグ'].dropna().astype(str))
    circle_slugs = {slug for slug in circle_slugs if slug.strip() and slug not in ['nan', 'None', '<NA>']}
    
    last_month_slugs = set(last_month_data['スラッグ'].dropna().astype(str))
    last_month_slugs = {slug for slug in last_month_slugs if slug.strip() and slug not in ['nan', 'None', '<NA>']}
    
    # 育児サークルデータにのみ存在するスラッグ
    only_in_circle = circle_slugs - last_month_slugs
    # 先月分データにのみ存在するスラッグ
    only_in_last_month = last_month_slugs - circle_slugs
    
    if only_in_circle:
        error_messages.append("\n### 先月分データに存在しないスラッグ:")
        for slug in only_in_circle:
            matching_rows = circle_data[circle_data['スラッグ'] == slug]
            if not matching_rows.empty:
                circle_name = matching_rows['サークル名'].iloc[0]
                error_messages.append(f"- スラッグ: {slug} (サークル名: {circle_name})")
    
    if only_in_last_month:
        error_messages.append("\n### 育児サークルデータに存在しないスラッグ:")
        for slug in only_in_last_month:
            matching_rows = last_month_data[last_month_data['スラッグ'] == slug]
            if not matching_rows.empty:
                circle_name = matching_rows['サークル名'].iloc[0]
                error_messages.append(f"- スラッグ: {slug} (サークル名: {circle_name})")
    
    if error_messages:
        st.error("""
        ### データの不一致が検出されました
        
        {}
        
        ※ スラッグの重複や不一致を修正してから再度実行してください。
        """.format('\n'.join(error_messages)))
        st.stop()

def add_account_columns(circle_data, last_month_data):
    """
    先月分のデータからアカウント情報を追加する
    
    Args:
        circle_data (pd.DataFrame): 育児サークルデータ
        last_month_data (pd.DataFrame): 先月分のデータ
    
    Returns:
        tuple: (処理後のデータフレーム, 処理内容のデータフレーム)
    """
    # アカウント関連列の追加
    account_columns = ['ｱｶｳﾝﾄ発行有無', 'ｱｶｳﾝﾄ発行年月', 'アカウント発行の登録用メールアドレス']
    
    try:
        # 先月分データのスラッグ列の重複チェック
        if 'スラッグ' not in last_month_data.columns:
            st.error("先月分データに「スラッグ」列が存在しません。")
            st.stop()
        
        # スラッグの重複チェック
        slug_duplicates = last_month_data[last_month_data['スラッグ'].duplicated()]['スラッグ'].unique()
        if len(slug_duplicates) > 0:
            error_message = "### 先月分データ内で重複しているスラッグが検出されました:\n"
            for slug in slug_duplicates:
                duplicate_rows = last_month_data[last_month_data['スラッグ'] == slug]
                error_message += f"- スラッグ: {slug}\n"
                for idx, row in duplicate_rows.iterrows():
                    circle_name = row.get('サークル名', '不明')
                    error_message += f"  - 行{idx+1}: {circle_name}\n"
            
            error_message += "\n※ 先月分データのスラッグ重複を修正してから再度実行してください。"
            st.error(error_message)
            st.stop()
        
        # 必要な列の存在チェック
        missing_columns = [col for col in account_columns if col not in last_month_data.columns]
        if missing_columns:
            st.error(f"先月分データに以下の列が存在しません: {', '.join(missing_columns)}")
            st.stop()
        
        # アカウント情報の追加
        for col in account_columns:
            # スラッグをキーとしてマッピング
            mapping_dict = last_month_data.set_index('スラッグ')[col].to_dict()
            circle_data[col] = circle_data['スラッグ'].map(mapping_dict)
            
    except Exception as e:
        st.error(f"""
        アカウント情報の追加中にエラーが発生しました。
        エラー内容: {str(e)}
        
        以下を確認してください：
        1. 先月分データのスラッグに重複がないこと
        2. 必要な列（{', '.join(account_columns)}）が先月分データに存在すること
        3. データの形式が正しいこと
        """)
        st.stop()
    
    # 処理内容のデータフレームを作成
    process_df = pd.DataFrame({
        '処理内容': ['先月分データからアカウント情報を追加'],
        '対象列': [', '.join(account_columns)]
    })
    
    return circle_data, process_df

def validate_csv_file(csv_file):
    """CSVファイルの検証を行う（セキュリティと品質を維持した最適化版）"""
    import time
    
    # 開始時間を記録
    start_time = time.time()
    debug_info = []
    timing_info = []
    
    # ファイルの内容を一度だけ読み込む（最大サイズを制限）
    file_read_start = time.time()
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    file_content = csv_file.read(MAX_FILE_SIZE)
    if len(file_content) == MAX_FILE_SIZE:
        raise ValueError("ファイルサイズが制限を超えています（最大10MB）")
    
    file_read_time = time.time() - file_read_start
    timing_info.append(f"ファイル読み込み: {file_read_time:.3f}秒")
    
    # chardetによるエンコーディング検出（処理時間短縮のため一時的にコメントアウト）
    # detected_enc, confidence = detect_encoding(file_content)
    # debug_info.append(f"chardetが検出したエンコーディング: {detected_enc} (信頼度: {confidence:.2f})")
    
    # 試行するエンコーディングの順序を決定（固定順序で高速化）
    # encodings = [detected_enc] if detected_enc else []
    # encodings.extend(['utf-8', 'shift-jis', 'cp932', 'euc-jp'])
    # encodings = list(dict.fromkeys(encodings))
    encodings = ['utf-8-sig', 'utf-8', 'shift-jis', 'cp932', 'euc-jp']
    debug_info.append("エンコーディング検出をスキップし、固定順序で試行します（UTF-8 BOM対応）")
    
    encoding_start = time.time()
    successful_encoding = None
    
    for encoding in encodings:
        try:
            encoding_try_start = time.time()
            debug_info.append(f"エンコーディング {encoding} で試行中...")
            
            # まず一部のデータでテスト（先頭1000バイト）
            sample_test_start = time.time()
            sample_size = min(1000, len(file_content))
            try:
                sample = file_content[:sample_size].decode(encoding)
            except UnicodeDecodeError:
                # サンプルテストで失敗した場合、エラー処理付きで再試行
                try:
                    sample = file_content[:sample_size].decode(encoding, errors='ignore')
                    debug_info.append(f"  → サンプルデコードでエラー文字を無視しました")
                except:
                    debug_info.append(f"  → サンプルデコードに失敗")
                    continue
            
            sample_test_time = time.time() - sample_test_start
            timing_info.append(f"サンプルテスト({encoding}): {sample_test_time:.3f}秒")
            
            if not sample.strip():
                debug_info.append(f"  → サンプルデータが空です")
                continue
            
            # 全体をデコード（エラー処理付き）
            full_decode_start = time.time()
            try:
                decoded_content = file_content.decode(encoding)
            except UnicodeDecodeError:
                # エラー文字を無視してデコード
                decoded_content = file_content.decode(encoding, errors='ignore')
                debug_info.append(f"  → 全体デコードでエラー文字を無視しました")
            
            full_decode_time = time.time() - full_decode_start
            timing_info.append(f"全体デコード({encoding}): {full_decode_time:.3f}秒")
            
            # CSVとしての基本検証
            csv_parse_start = time.time()
            df = pd.read_csv(io.StringIO(decoded_content))
            csv_parse_time = time.time() - csv_parse_start
            timing_info.append(f"CSV解析({encoding}): {csv_parse_time:.3f}秒")
            
            # データ品質の検証
            validation_start = time.time()

            # CSVファイルとしての基本的な構造確認のみ
            if df.empty:
                debug_info.append(f"  → データが空です")
                continue
            if len(df.columns) == 0:
                debug_info.append(f"  → 列が存在しません")
                continue
            
            validation_time = time.time() - validation_start
            timing_info.append(f"データ検証({encoding}): {validation_time:.3f}秒")
            
            encoding_try_time = time.time() - encoding_try_start
            timing_info.append(f"エンコーディング試行完了({encoding}): {encoding_try_time:.3f}秒")
            
            debug_info.append(f"  → 正常に読み込めました")
            successful_encoding = encoding
            break
            
        except UnicodeDecodeError:
            debug_info.append(f"  → デコードエラー")
            continue
        except pd.errors.EmptyDataError:
            debug_info.append(f"  → 空のCSVファイル")
            raise ValueError("CSVファイルが空です")
        except ValueError as e:
            # 検証エラーは上位に伝播
            raise e
        except Exception as e:
            debug_info.append(f"  → その他のエラー: {str(e)}")
            continue
    
    if successful_encoding is None:
        error_msg = "CSVファイルのエンコーディングを認識できません。以下のいずれかの形式で保存してください：UTF-8、Shift-JIS、CP932、EUC-JP"
        if st.session_state.get('debug_mode', False):
            error_msg += "\n\nデバッグ情報:\n" + "\n".join(debug_info)
        raise ValueError(error_msg)
    
    encoding_total_time = time.time() - encoding_start
    timing_info.append(f"エンコーディング処理合計: {encoding_total_time:.3f}秒")
    
    total_time = time.time() - start_time
    timing_info.append(f"全体処理時間: {total_time:.3f}秒")
    
    # デバッグモード時に処理時間を表示
    if st.session_state.get('debug_mode', False):
        st.write("**⏱️ CSVファイル読み込み処理時間:**")
        for timing in timing_info:
            st.write(f"  - {timing}")
    
    return df, successful_encoding, debug_info

def copy_cell_format(source_cell, target_cell):
    """セルの書式をコピーする"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def validate_excel_file(excel_file):
    """先月分のExcelファイルの検証と読み込みを行う
    
    Args:
        excel_file: アップロードされたExcelファイル
    
    Returns:
        pd.DataFrame: 読み込んだデータフレーム
    """
    import time
    
    # 開始時間を記録
    start_time = time.time()
    timing_info = []
    
    try:
        # Excelファイルを読み込む（2,3行目をスキップ）
        excel_read_start = time.time()
        df = pd.read_excel(excel_file, skiprows=[1,2])
        excel_read_time = time.time() - excel_read_start
        timing_info.append(f"Excelファイル読み込み: {excel_read_time:.3f}秒")
        
        # 基本的な検証
        validation_start = time.time()
        if df.empty:
            raise ValueError("Excelファイルにデータが存在しません")
            
        if len(df.columns) == 0:
            raise ValueError("Excelファイルに列が存在しません")
        
        # ヘッダーの存在確認
        if df.columns.isna().any():
            raise ValueError("ヘッダー行に空の列名が存在します")
        
        validation_time = time.time() - validation_start
        timing_info.append(f"基本検証: {validation_time:.3f}秒")
        
        total_time = time.time() - start_time
        timing_info.append(f"全体処理時間: {total_time:.3f}秒")
        
        # デバッグモード時に処理時間を表示
        if st.session_state.get('debug_mode', False):
            st.write("**⏱️ Excelファイル読み込み処理時間:**")
            for timing in timing_info:
                st.write(f"  - {timing}")
        
        return df
        
    except pd.errors.EmptyDataError:
        raise ValueError("Excelファイルが空です")
    except Exception as e:
        raise ValueError(f"Excelファイルの読み込み中にエラーが発生しました: {str(e)}")

def hide_columns(worksheet):
    """特定の列を非表示にする
    
    Args:
        worksheet: 対象のワークシート
    """
    # 非表示にする列名のリスト
    columns_to_hide = [
        'スラッグ',
        'ステータス',
        '参加者の条件(妊娠後半)',
        '参加者の条件(出産)',
        '参加者の条件(1歳後半)',
        '参加者の条件(2歳後半)',
        '申込方法備考',
        '活動日_営業時間ラベル',
        '活動日_営業曜日ラベル',
        '代表者',
        '団体名'
    ]
    
    # ヘッダー行から列のインデックスを取得
    header_row = 1  # ヘッダーは1行目にある
    for column in worksheet.iter_cols(min_row=header_row, max_row=header_row):
        if column[0].value in columns_to_hide:
            col_letter = get_column_letter(column[0].column)
            worksheet.column_dimensions[col_letter].hidden = True

def add_borders(worksheet, start_row, end_row, start_col, end_col):
    """データ範囲に枠線を追加する
    
    Args:
        worksheet: 対象のワークシート
        start_row: 開始行（1始まり）
        end_row: 終了行（1始まり）
        start_col: 開始列（1始まり）
        end_col: 終了列（1始まり）
    """
    # 枠線のスタイルを定義
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 指定範囲の各セルに枠線を設定
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border

def find_data_range(worksheet):
    """データが存在する範囲を特定する
    
    Args:
        worksheet: 対象のワークシート
    
    Returns:
        tuple: (最終行, 最終列)
    """
    max_row = 1
    max_col = 1
    
    # 最終行を特定
    for row in worksheet.iter_rows():
        if any(cell.value is not None for cell in row):
            max_row = row[0].row
    
    # 最終列を特定
    for col in worksheet.iter_cols():
        if any(cell.value is not None for cell in col):
            max_col = col[0].column
    
    return max_row, max_col

def set_row_height_and_format(worksheet, start_row, end_row, height=20):
    """行の高さを設定し、セルの書式を設定する
    
    Args:
        worksheet: 対象のワークシート
        start_row: 開始行（1始まり）
        end_row: 終了行（1始まり）
        height: 行の高さ（デフォルト: 20）
    """
    # セルの書式設定（折り返し有効、左揃え）
    alignment = Alignment(
        wrap_text=True,  # 折り返し
        horizontal='left',  # 左揃え
        vertical='center'  # 縦方向は中央揃え
    )
    
    # フォント設定
    font = Font(
        name='メイリオ',  # フォント名
        size=12,         # フォントサイズ
    )
    
    # 指定範囲の各行に対して設定
    for row in range(start_row, end_row + 1):
        # 行の高さを設定
        worksheet.row_dimensions[row].height = height
        
        # その行の各セルの書式を設定
        for cell in worksheet[row]:
            cell.alignment = alignment
            cell.font = font

def setup_conditional_formatting(worksheet):
    """条件付き書式を設定する
    
    Args:
        worksheet: 対象のワークシート
    """
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import PatternFill
    from openpyxl.styles.differential import DifferentialStyle
    
    # 色の定義
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
    
    # スタイルの定義
    red_style = DifferentialStyle(fill=red_fill)
    yellow_style = DifferentialStyle(fill=yellow_fill)
    green_style = DifferentialStyle(fill=green_fill)
    
    # 条件付き書式のリスト
    conditional_rules = [
        {
            'name': 'スラッグの差分検出',
            'description': 'スラッグが空、または入力されているがoriginalに見つからないものを検出',
            'formula': 'OR($B1="",ISERROR(MATCH($B1,INDIRECT("original!B1:B1048576"),0)))',
            'range': 'B1:B1048576',
            'style': red_style
        },
        # 追加の条件付き書式はここに追加
                 {
             'name': '変更箇所の検出',
             'description': '同じスラッグを持つ行の同じ列のセルを比較　⇒ 該当するセルだけ黄色く着色',
             'formula': 'A1<>INDIRECT("original!"&ADDRESS(MATCH($B1,INDIRECT("original!B1:B1048576"),0),COLUMN(),4,1))',
             'range': 'A1:ZZ1048576',
             'style': yellow_style
         },
         {
             'name': '追加行の検出',
             'description': '入力されているスラッグがoriginalに見つからないおよびサークル名がoriginalに見つからない',
             'formula': 'OR(ISERROR(MATCH($B1,INDIRECT("original!B1:B1048576"),0)),ISERROR(MATCH($C1,INDIRECT("original!C1:C1048576"),0)))',
             'range': 'A1:ZZ1048576',
             'style': green_style
         }
    ]
    
    # 条件付き書式を適用
    for rule_config in conditional_rules:
        rule = Rule(
            type="expression",
            formula=[rule_config['formula']],
            stopIfTrue=True,
            dxf=rule_config['style']
        )
        worksheet.conditional_formatting.add(rule_config['range'], rule)
        
        # デバッグモード時に設定内容を出力
        if st.session_state.get('debug_mode', False):
            st.info(f"条件付き書式を設定: {rule_config['name']} - {rule_config['description']}")

def process_files(circle_data, facility_data=None, last_month_data=None):
    """Pandasを使用したファイル処理"""
    start_time = time.time() if st.session_state.get('debug_mode', False) else None
    
    # 処理内容を記録するデータフレームを作成
    process_df = pd.DataFrame(columns=['処理内容', '対象列'])
    
    # 0/1の値を持つ列の変換処理
    circle_data, binary_process_df = process_binary_columns(circle_data)
    if not binary_process_df.empty:
        process_df = pd.concat([process_df, binary_process_df], ignore_index=True)
    
    # 場所列の追加
    circle_data, location_process_df = add_location_column(circle_data,facility_data)
    process_df = pd.concat([process_df, location_process_df], ignore_index=True)
    
    # アカウント情報の追加
    circle_data, account_process_df = add_account_columns(circle_data, last_month_data)
    process_df = pd.concat([process_df, account_process_df], ignore_index=True)
    
    # 処理内容と処理後データの表示は外部で行う（UIの流れを改善するため）
    # この関数からは処理内容データも返すように変更
    
    # ファイルを保存
    output = io.BytesIO()
    template_wb = load_workbook(TEMPLATE_FILE)
    template_ws = template_wb.active
    
    # シート名を'original'に変更
    template_ws.title = 'original'
    
    # テンプレートの内容をそのままコピー
    template_wb.save(output)
    output.seek(0)
    
    # 保存したファイルを再度開く
    wb = load_workbook(output)
    original_ws = wb['original']
    
    # CSVデータを書き込む（ヘッダー行を除く）
    if len(circle_data) > 0:  # データが存在する場合のみ処理
        # CSVの列数がテンプレートの列数を超えていないかチェック
        if len(circle_data.columns) > template_ws.max_column:
            raise ValueError(f"CSVファイルの列数（{len(circle_data.columns)}列）がテンプレートの列数（{template_ws.max_column}列）を超えています。")
        
        # データを一括で書き込む
        data_values = circle_data.values
        for row_idx, row in enumerate(data_values, start=4):  # 4行目から開始
            for col_idx, value in enumerate(row, start=1):
                cell = original_ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                # テンプレートの同じ位置のセルから書式をコピー
                template_cell = template_ws.cell(row=row_idx, column=col_idx)
                copy_cell_format(template_cell, cell)
        
        # データが存在する範囲を特定
        max_row, max_col = find_data_range(original_ws)
        
        # データ部分に枠線を追加（1行目から最終行まで）
        add_borders(original_ws, 1, max_row, 1, max_col)
        
        # 行の高さとセル書式を設定（4行目からデータ最終行まで）
        set_row_height_and_format(original_ws, 4, max_row)
    
    # 特定の列を非表示にする
    hide_columns(original_ws)
    
    # シートをコピーして'circle_info'シートを作成
    circle_info_ws = wb.copy_worksheet(original_ws)
    circle_info_ws.title = 'circle_info'
    
    # originalシートを非表示にする
    original_ws.sheet_state = 'hidden'
    
    # 条件付き書式の設定
    setup_conditional_formatting(circle_info_ws)
    
    # シートのグループを解除
    for ws in wb.worksheets:
        ws.sheet_view.tabSelected = False
    
    # アクティブシートを明示的に設定（circle_infoシートをアクティブに）
    wb.active = circle_info_ws
    
    # ファイルを保存
    output.seek(0)
    wb.save(output)
    
    processing_time = time.time() - start_time if st.session_state.get('debug_mode', False) else None
    
    output.seek(0)
    return output, processing_time, process_df, circle_data

def get_openai_client():
    """OpenAI APIクライアントを取得する"""
    try:
        api_key = st.secrets["openai"]["api_key"]
        if api_key == "YOUR_OPENAI_API_KEY_HERE":
            return None
        
        client = OpenAI(
            api_key=api_key
        )
        return client
    except Exception as e:
        st.error(f"OpenAI API設定エラー: {str(e)}")
        return None

def get_robot_icon():
    """ロボットアイコンを取得する（画像ファイルまたは絵文字）"""
    # 複数の画像ファイルパスを試行（アップロードされた画像を優先）
    possible_paths = [
        "img/bot.png",  # アップロードされたキャラクター画像
        "robot_icon.png",
        "images/robot.png", 
        "assets/robot.png",
        "robot_icon.jpg",
        "images/robot.jpg",
        "assets/robot.jpg",
        "robot_icon.gif",
        "images/robot.gif",
        "assets/robot.gif"
    ]
    
    for robot_image_path in possible_paths:
        if os.path.exists(robot_image_path):
            try:
                # Base64エンコーディングで画像を埋め込む
                import base64
                with open(robot_image_path, "rb") as img_file:
                    img_data = base64.b64encode(img_file.read()).decode()
                    
                    # ファイル拡張子に応じてMIMEタイプを決定
                    if robot_image_path.lower().endswith('.png'):
                        mime_type = "image/png"
                    elif robot_image_path.lower().endswith('.jpg') or robot_image_path.lower().endswith('.jpeg'):
                        mime_type = "image/jpeg"
                    elif robot_image_path.lower().endswith('.gif'):
                        mime_type = "image/gif"
                    else:
                        mime_type = "image/png"  # デフォルト
                    
                    # アップロードされたキャラクター画像の場合は少し大きく表示
                    if "bot.png" in robot_image_path:
                        return f'<img src="data:{mime_type};base64,{img_data}" style="width: 32px; height: 32px; border-radius: 50%; object-fit: cover;">'
                    else:
                        return f'<img src="data:{mime_type};base64,{img_data}" style="width: 24px; height: 24px; border-radius: 50%; object-fit: cover;">'
            except Exception as e:
                # 画像読み込みエラーの場合は次のパスを試行
                continue
    
    # 画像ファイルが存在しない場合は絵文字を使用
    return "🤖"

def get_user_icon():
    """ユーザーアイコンを取得する（画像ファイルまたは絵文字）"""
    # 複数の画像ファイルパスを試行
    possible_paths = [
        "user_icon.png",
        "images/user.png", 
        "assets/user.png",
        "user_icon.jpg",
        "images/user.jpg",
        "assets/user.jpg"
    ]
    
    for user_image_path in possible_paths:
        if os.path.exists(user_image_path):
            try:
                # Base64エンコーディングで画像を埋め込む
                import base64
                with open(user_image_path, "rb") as img_file:
                    img_data = base64.b64encode(img_file.read()).decode()
                    
                    # ファイル拡張子に応じてMIMEタイプを決定
                    if user_image_path.lower().endswith('.png'):
                        mime_type = "image/png"
                    elif user_image_path.lower().endswith('.jpg') or user_image_path.lower().endswith('.jpeg'):
                        mime_type = "image/jpeg"
                    else:
                        mime_type = "image/png"  # デフォルト
                    
                    return f'<img src="data:{mime_type};base64,{img_data}" style="width: 24px; height: 24px; border-radius: 50%; object-fit: cover;">'
            except Exception as e:
                # 画像読み込みエラーの場合は次のパスを試行
                continue
    
    # 画像ファイルが存在しない場合は絵文字を使用
    return "👤"

def get_codebase_context():
    """外部ファイルから詳細仕様書を読み込む"""
    import time
    import streamlit as st
    
    # 開始時間を記録
    start_time = time.time()
    timing_info = []
    
    try:
        # ファイルパス取得の時間測定
        path_start = time.time()
        spec_file_path = os.path.join(os.path.dirname(__file__), 'app_specification.md')
        path_time = time.time() - path_start
        timing_info.append(f"ファイルパス取得: {path_time:.3f}秒")
        
        # ファイル存在チェックの時間測定
        check_start = time.time()
        file_exists = os.path.exists(spec_file_path)
        check_time = time.time() - check_start
        timing_info.append(f"ファイル存在チェック: {check_time:.3f}秒")
        
        # ファイルが存在しない場合のフォールバック
        if not file_exists:
            fallback_content = """
            # 育児サークル情報処理ツール - 基本情報
            
            仕様書ファイル (app_specification.md) が見つかりません。
            基本的な使い方については、サイドバーの使い方説明を参照してください。
            
            ## 主な機能
            1. データ修正用エクセル作成
            2. インポートデータ作成（16項目検証）
            3. AIチャット機能
            """
            
            total_time = time.time() - start_time
            timing_info.append(f"フォールバック処理: {total_time:.3f}秒")
            
            # デバッグモード時に処理時間を表示
            if st.session_state.get('debug_mode', False):
                st.sidebar.markdown("**⏱️ 仕様書読み込み処理時間（フォールバック）:**")
                for timing in timing_info:
                    st.sidebar.text(f"  - {timing}")
            
            return fallback_content
        
        # ファイルから仕様書を読み込み
        file_read_start = time.time()
        with open(spec_file_path, 'r', encoding='utf-8') as f:
            context = f.read()
        file_read_time = time.time() - file_read_start
        timing_info.append(f"ファイル読み込み: {file_read_time:.3f}秒")
        
        # 全体処理時間
        total_time = time.time() - start_time
        timing_info.append(f"全体処理時間: {total_time:.3f}秒")
        
        # デバッグモード時に処理時間を表示
        if st.session_state.get('debug_mode', False):
            st.sidebar.markdown("**⏱️ 仕様書読み込み処理時間:**")
            for timing in timing_info:
                st.sidebar.text(f"  - {timing}")
            st.sidebar.info(f"仕様書サイズ: {len(context):,} 文字")
        
        return context
        
    except Exception as e:
        error_time = time.time() - start_time
        timing_info.append(f"エラー発生まで: {error_time:.3f}秒")
        
        # デバッグモード時にエラー時の処理時間も表示
        if st.session_state.get('debug_mode', False):
            st.sidebar.markdown("**⏱️ 仕様書読み込み処理時間（エラー）:**")
            for timing in timing_info:
                st.sidebar.text(f"  - {timing}")
            st.sidebar.error(f"ファイル読み込みエラー: {str(e)}")
        
        # エラー時のフォールバック
        return f"""
        # 育児サークル情報処理ツール - エラー
        
        仕様書の読み込み中にエラーが発生しました: {str(e)}
        
        ## 基本機能
        - データ修正用エクセル作成
        - インポートデータ作成
        - 16項目のデータ検証
        - AIチャット機能
        """

def chat_with_openai(client, message, context):
    """OpenAI APIを使用してチャット応答を生成する"""
    import time
    import streamlit as st
    
    # 開始時間を記録
    start_time = time.time()
    timing_info = []
    
    try:
        # メッセージ準備の時間測定
        message_prep_start = time.time()
        messages = [
            {
                "role": "system",
                "content": f"""あなたは育児サークル情報処理ツールの専門サポートエージェントです。
                以下のコードベース情報を参考に、ユーザーの質問に日本語で回答してください。
                
                {context}
                
                回答は簡潔で分かりやすく、具体的な手順を含めてください。
                技術的な詳細よりも、実際の使用方法に焦点を当ててください。"""
            },
            {
                "role": "user",
                "content": message
            }
        ]
        message_prep_time = time.time() - message_prep_start
        timing_info.append(f"メッセージ準備: {message_prep_time:.3f}秒")
        
        # OpenAI API呼び出しの時間測定
        api_call_start = time.time()
        response = client.chat.completions.create(
            model="gpt-4.1-mini-2025-04-14",
            messages=messages,
            max_tokens=1500,
            temperature=0.1
        )
        api_call_time = time.time() - api_call_start
        timing_info.append(f"OpenAI API呼び出し: {api_call_time:.3f}秒")
        
        # レスポンス処理の時間測定
        response_process_start = time.time()
        response_content = response.choices[0].message.content
        response_process_time = time.time() - response_process_start
        timing_info.append(f"レスポンス処理: {response_process_time:.3f}秒")
        
        # 全体処理時間
        total_time = time.time() - start_time
        timing_info.append(f"全体処理時間: {total_time:.3f}秒")
        
        # デバッグモード時に処理時間を表示
        if st.session_state.get('debug_mode', False):
            st.sidebar.markdown("**⏱️ サポートチャット処理時間:**")
            for timing in timing_info:
                st.sidebar.text(f"  - {timing}")
        
        return response_content
        
    except Exception as e:
        error_time = time.time() - start_time
        timing_info.append(f"エラー発生まで: {error_time:.3f}秒")
        
        # デバッグモード時にエラー時の処理時間も表示
        if st.session_state.get('debug_mode', False):
            st.sidebar.markdown("**⏱️ サポートチャット処理時間（エラー）:**")
            for timing in timing_info:
                st.sidebar.text(f"  - {timing}")
            st.sidebar.error(f"API呼び出しエラー: {str(e)}")
        
        return f"申し訳ございません。エラーが発生しました: {str(e)}"

def show_sidebar_chat():
    """サイドバーにチャット機能を表示する"""
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 💬 サポートチャット")
    
    # APIキーの確認
    client = get_openai_client()
    if client is None:
        st.sidebar.warning("⚠️ OpenAI APIキーが設定されていません。")
        st.sidebar.markdown("""
        **設定方法:**
        1. `.streamlit/secrets.toml` を編集
        2. `[openai]` セクションに `api_key = "your_actual_api_key"` を設定
        3. ツールを再起動
        """)
        return
    
    # チャット履歴の初期化
    if 'chat_history' not in st.session_state:
        st.session_state.chat_history = []
    
    # チャット履歴の表示（古い順に表示）
    if st.session_state.chat_history:
        # チャット履歴を上から古い順に表示（ChatGPT/Cursor風）
        chat_container = st.sidebar.container()
        
        with chat_container:
            # 最新の5件を表示（サイドバーのスペース制限のため）
            recent_chats = st.session_state.chat_history[-5:]
            
            for i, chat in enumerate(recent_chats):
                # ユーザーの質問（ユーザーアイコン付き）
                user_icon = get_user_icon()
                
                import html
                escaped_user = html.escape(chat['user'])
                st.sidebar.markdown(f"""
                <div style="background-color: #f0f2f6; padding: 8px; border-radius: 8px; margin-bottom: 5px; display: flex; align-items: flex-start;">
                    <div style="margin-right: 8px; font-size: 20px;">{user_icon}</div>
                    <div style="flex: 1;">
                        <small style="color: #666;">🕐 {chat['timestamp']}</small><br>
                        <strong>質問:</strong> {escaped_user}
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
                # AIの回答（ロボットアイコン付き）
                robot_icon = get_robot_icon()
                
                escaped_assistant = html.escape(chat['assistant'])
                st.sidebar.markdown(f"""
                <div style="background-color: #e8f4f8; padding: 8px; border-radius: 8px; margin-bottom: 10px; display: flex; align-items: flex-start;">
                    <div style="margin-right: 8px; font-size: 20px;">{robot_icon}</div>
                    <div style="flex: 1;">
                        {escaped_assistant}
                    </div>
                </div>
                """, unsafe_allow_html=True)
    
    # チャット入力（履歴の下に配置）
    user_input = st.sidebar.text_area(
        "質問を入力してください:",
        height=80,
        placeholder="例: エラーが出た時はどうすればいいですか？",
        key="chat_input"
    )
    
    # 送信ボタン
    if st.sidebar.button("📤 送信", key="chat_send", use_container_width=True):
        if user_input.strip():
            import time
            
            # 全体処理時間の測定開始
            overall_start = time.time()
            
            # コードベースの文脈を取得
            context_start = time.time()
            context = get_codebase_context()
            context_time = time.time() - context_start
            
            # チャット応答を生成
            # Streamlit 1.30.0以降でst.sidebar.spinner()が利用可能
            try:
                with st.sidebar.spinner("回答を生成中..."):
                    chat_start = time.time()
                    response = chat_with_openai(client, user_input, context)
                    chat_time = time.time() - chat_start
            except AttributeError:
                # 古いStreamlitバージョンの場合は通常のspinnerを使用
                with st.spinner("回答を生成中..."):
                    chat_start = time.time()
                    response = chat_with_openai(client, user_input, context)
                    chat_time = time.time() - chat_start
            
            # チャット履歴に追加
            history_start = time.time()
            st.session_state.chat_history.append({
                "user": user_input,
                "assistant": response,
                "timestamp": datetime.datetime.now().strftime("%H:%M:%S")
            })
            history_time = time.time() - history_start
            
            # 全体処理時間
            overall_time = time.time() - overall_start
            
            # デバッグモード時に全体の処理時間を表示
            if st.session_state.get('debug_mode', False):
                st.sidebar.markdown("**⏱️ サポートチャット全体処理時間:**")
                st.sidebar.text(f"  - コンテキスト取得: {context_time:.3f}秒")
                st.sidebar.text(f"  - チャット応答生成: {chat_time:.3f}秒")
                st.sidebar.text(f"  - 履歴追加: {history_time:.3f}秒")
                st.sidebar.text(f"  - 全体処理時間: {overall_time:.3f}秒")
                st.sidebar.markdown("---")
            
            # 入力をクリア
            st.rerun()
    
    # 履歴クリアボタン（送信ボタンの下に配置）
    if st.session_state.chat_history:
        if st.sidebar.button("🗑️ 履歴をクリア", key="chat_clear"):
            st.session_state.chat_history = []
            st.rerun()
    


def initialize_session_state():
    """セッション状態の初期化"""
    if 'debug_mode' not in st.session_state:
        st.session_state.debug_mode = False
    
    # 機能選択の前回状態を追加
    if 'previous_function' not in st.session_state:
        st.session_state.previous_function = None
    
    # インポートデータ作成用のセッション状態
    if 'validation_completed' not in st.session_state:
        st.session_state.validation_completed = False
    if 'validated_data' not in st.session_state:
        st.session_state.validated_data = None
    if 'import_data_created' not in st.session_state:
        st.session_state.import_data_created = False
    if 'import_files' not in st.session_state:
        st.session_state.import_files = None
    if 'formatted_data' not in st.session_state:
        st.session_state.formatted_data = None
    
    # セッション状態の変化を追跡するためのログ
    if 'session_log' not in st.session_state:
        st.session_state.session_log = []
    
    # アップロードされたファイルの追跡用
    if 'uploaded_files_hash' not in st.session_state:
        st.session_state.uploaded_files_hash = {
            'excel': None,
            'facility': None,
            'user': None
        }

def log_session_state_change(action, details=None):
    """セッション状態の変化をログに記録"""
    import datetime
    
    if 'session_log' not in st.session_state:
        st.session_state.session_log = []
    
    log_entry = {
        'timestamp': datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3],
        'action': action,
        'details': details or {},
        'session_state': {
            'validation_completed': st.session_state.get('validation_completed', False),
            'validated_data': st.session_state.get('validated_data') is not None,
            'import_data_created': st.session_state.get('import_data_created', False),
            'import_files': st.session_state.get('import_files') is not None,
            'formatted_data': st.session_state.get('formatted_data') is not None,
        }
    }
    
    st.session_state.session_log.append(log_entry)
    
    # ログの最大数を制限（メモリ使用量を抑制）
    if len(st.session_state.session_log) > 50:
        st.session_state.session_log = st.session_state.session_log[-50:]

def show_session_state_debug():
    """セッション状態のデバッグ情報を表示"""
    if not st.session_state.get('debug_mode', False):
        return
    
    with st.expander("🔍 セッション状態デバッグ情報", expanded=False):
        # 現在のセッション状態
        st.subheader("現在のセッション状態")
        current_state = {
            'validation_completed': st.session_state.get('validation_completed', False),
            'validated_data': st.session_state.get('validated_data') is not None,
            'import_data_created': st.session_state.get('import_data_created', False),
            'import_files': st.session_state.get('import_files') is not None,
            'formatted_data': st.session_state.get('formatted_data') is not None,
        }
        
        col1, col2 = st.columns(2)
        with col1:
            for key, value in current_state.items():
                status_icon = "✅" if value else "❌"
                st.write(f"{status_icon} {key}: {value}")
        
        with col2:
            if st.session_state.get('validated_data') is not None:
                st.write(f"📊 validated_data 行数: {len(st.session_state.validated_data)}")
            if st.session_state.get('import_files') is not None:
                st.write(f"📁 import_files 数: {len(st.session_state.import_files)}")
            if st.session_state.get('formatted_data') is not None:
                st.write(f"📋 formatted_data 行数: {len(st.session_state.formatted_data)}")
        
        # セッション状態の変化ログ
        if st.session_state.get('session_log'):
            st.subheader("セッション状態変化ログ")
            
            # ログをテーブル形式で表示
            log_data = []
            for log_entry in reversed(st.session_state.session_log[-10:]):  # 最新10件
                log_data.append({
                    '時刻': log_entry['timestamp'],
                    'アクション': log_entry['action'],
                    '検証完了': "✅" if log_entry['session_state']['validation_completed'] else "❌",
                    'データ作成完了': "✅" if log_entry['session_state']['import_data_created'] else "❌",
                    'ファイル存在': "✅" if log_entry['session_state']['import_files'] else "❌",
                    '詳細': str(log_entry['details']) if log_entry['details'] else ""
                })
            
            if log_data:
                st.dataframe(pd.DataFrame(log_data), use_container_width=True)
            
            # ログクリアボタン
            if st.button("ログをクリア", key="clear_session_log"):
                st.session_state.session_log = []
                # st.rerun()は削除 - セッションリセットを防ぐため、次回の自動リロードで反映される
        
        # ユーザーCSV作成デバッグ情報
        if hasattr(st.session_state, 'user_csv_debug_info') and st.session_state.user_csv_debug_info:
            st.subheader("ユーザー新規追加CSV作成デバッグ情報")
            debug_info = st.session_state.user_csv_debug_info
            
            col1, col2 = st.columns(2)
            with col1:
                st.write(f"📊 全データ行数: {debug_info['total_rows']}")
                st.write(f"✅ アカウント発行有無が○の行数: {debug_info['account_issued_count']}")
            
            with col2:
                st.write(f"📧 メールアドレス記載の行数: {debug_info['email_filled_count']}")
                st.write(f"🎯 両方の条件を満たす行数: {debug_info['new_accounts_count']}")
            
            # サンプルデータまたは原因調査の表示
            if debug_info['new_accounts_count'] > 0:
                st.write("✅ **条件を満たすデータのサンプル:**")
                if debug_info['new_accounts_sample'] is not None:
                    st.dataframe(debug_info['new_accounts_sample'], use_container_width=True)
            else:
                st.write("❌ **条件を満たさない理由の調査:**")
                
                if debug_info['account_values'] is not None:
                    st.write("**アカウント発行有無列の値の分布:**")
                    st.write(debug_info['account_values'])
                

            
            # デバッグ情報クリアボタン
            if st.button("ユーザーCSVデバッグ情報をクリア", key="clear_user_csv_debug"):
                del st.session_state.user_csv_debug_info
                # st.rerun()は削除 - セッションリセットを防ぐため、次回の自動リロードで反映される

def reset_import_session_state():
    """インポートデータ作成関連のセッション状態をリセット"""
    log_session_state_change("reset_import_session_state", {
        'before_validation_completed': st.session_state.get('validation_completed', False),
        'before_import_data_created': st.session_state.get('import_data_created', False)
    })
    
    st.session_state.validation_completed = False
    st.session_state.validated_data = None
    st.session_state.import_data_created = False
    st.session_state.import_files = None
    st.session_state.formatted_data = None
    st.session_state.balloons_shown = False  # バルーン表示フラグもリセット
    
    # 警告メッセージもクリア
    if 'account_date_warning' in st.session_state:
        del st.session_state.account_date_warning
    
    # ユーザー作成警告もクリア
    if 'user_creation_warning' in st.session_state:
        del st.session_state.user_creation_warning
    
    # ユーザー修正警告もクリア
    if 'user_modification_warning' in st.session_state:
        del st.session_state.user_modification_warning
    
    # ユーザー修正情報もクリア
    if 'user_modification_details' in st.session_state:
        del st.session_state.user_modification_details
    
    # 統合ユーザー情報もクリア
    if 'user_comprehensive_details' in st.session_state:
        del st.session_state.user_comprehensive_details
    
    # すでに発行済みユーザー情報もクリア
    if 'already_issued_users' in st.session_state:
        del st.session_state.already_issued_users

def check_file_changed(file, file_type):
    """ファイルが変更されたかチェックし、変更された場合のみセッション状態をリセット
    
    Args:
        file: アップロードされたファイル
        file_type: ファイルタイプ ('excel', 'facility', 'user')
    
    Returns:
        bool: ファイルが変更された場合True
    """
    if file is None:
        return False
    
    # ファイルの内容からハッシュを生成
    file_content = file.read()
    file.seek(0)  # ファイルポインタを先頭に戻す
    file_hash = hashlib.md5(file_content).hexdigest()
    
    # 前回のハッシュと比較
    previous_hash = st.session_state.uploaded_files_hash.get(file_type)
    
    if previous_hash != file_hash:
        # ファイルが変更された場合
        st.session_state.uploaded_files_hash[file_type] = file_hash
        log_session_state_change(f"{file_type}_file_changed", {
            'filename': file.name,
            'previous_hash': previous_hash,
            'new_hash': file_hash
        })
        return True
    else:
        # ファイルが変更されていない場合
        log_session_state_change(f"{file_type}_file_unchanged", {
            'filename': file.name,
            'hash': file_hash
        })
        return False

def validate_order_column(df):
    """順番列の値を検証する
    
    Args:
        df (pd.DataFrame): 検証対象のデータフレーム
    
    Raises:
        ValueError: 順番列に数値以外の値が含まれている場合
    """
    if '順番' not in df.columns:
        return
    
    # 数値以外の値を含む行を検出
    non_numeric_rows = df[pd.to_numeric(df['順番'], errors='coerce').isna()]
    if not non_numeric_rows.empty:
        error_message = ["### エラー: 「順番」列に数値以外の値が含まれています"]
        for _, row in non_numeric_rows.iterrows():
            error_message.append(f"- サークル名: {row['サークル名']}")
            error_message.append(f"  - スラッグ: {row['スラッグ']}")
            error_message.append(f"  - 順番: {row['順番']}")
        
        raise ValueError("\n".join(error_message))
    
    # 1未満の値を含む行を検出
    invalid_rows = df[pd.to_numeric(df['順番']) < 1]
    if not invalid_rows.empty:
        warning_message = ["### 警告: 「順番」列に1未満の値が含まれている行があります"]
        for _, row in invalid_rows.iterrows():
            warning_message.append(f"- サークル名: {row['サークル名']}")
            warning_message.append(f"  - スラッグ: {row['スラッグ']}")
            warning_message.append(f"  - 順番: {row['順番']}")
        
        st.warning("\n".join(warning_message))

def show_excel_creation_page():
    """データ修正用エクセル作成ページの表示"""
    # st.header("データ修正用エクセル作成", divider='orange')
    
    if st.session_state.debug_mode:
        st.write("デバッグモードが有効です")
    
    st.header('STEP1：ファイルのアップロード', divider='orange')
    
    # 育児サークルCSVファイルのアップロード
    st.write("**1. 育児サークル情報（csv）をアップロードしてください。**")
    csv_file = st.file_uploader("例: kitakyushu-city_circle_info_00000000000.csv", type=['csv'])
    if csv_file:
        try:
            # CSVファイルの検証と読み込み
            import time
            overall_start = time.time()
            circle_data, encoding, debug_info = validate_csv_file(csv_file)
            
            # 順番列の検証（検証の必要性について確認中。必要であればコメントアウト解除）
            # validate_order_column(circle_data)
            
            overall_time = time.time() - overall_start
            
            st.success("育児サークルCSVファイルが正常に読み込まれました")
            
            # デバッグモード時に追加情報を表示
            if st.session_state.get('debug_mode', False):
                st.write(f"**📊 データ情報:** 行数: {len(circle_data)}, 列数: {len(circle_data.columns)}")
                st.write(f"**🔤 エンコーディング:** {encoding}")
                st.write(f"**⏱️ 全体処理時間:** {overall_time:.3f}秒")
            
            with st.expander("育児サークルデータを確認する"):
                st.dataframe(circle_data, use_container_width=True)
        except ValueError as e:
            st.error(f"育児サークルCSVファイルのエラー: {str(e)}")
        except Exception as e:
            st.error(f"育児サークルCSVファイルの予期せぬエラー: {str(e)}")
    
    # 施設情報CSVファイルのアップロード
    st.write("**2. 施設情報（csv）をアップロードしてください。**")
    facility_csv_file = st.file_uploader("例: kitakyushu-city_facility_00000000000.csv", type=['csv'])
    if facility_csv_file:
        try:
            # 施設情報CSVファイルの検証と読み込み（専用の検証関数を使用）
            import time
            overall_start = time.time()
            facility_data, facility_encoding, facility_debug_info = validate_facility_csv_file(facility_csv_file)
            overall_time = time.time() - overall_start
            
            st.success("施設情報CSVファイルが正常に読み込まれました")
            
            # デバッグモード時に追加情報を表示
            if st.session_state.get('debug_mode', False):
                st.write(f"**📊 データ情報:** 行数: {len(facility_data)}, 列数: {len(facility_data.columns)}")
                st.write(f"**🔤 エンコーディング:** {facility_encoding}")
                st.write(f"**⏱️ 全体処理時間:** {overall_time:.3f}秒")
            
            with st.expander("施設情報データを確認する"):
                st.dataframe(facility_data, use_container_width=True)
        except ValueError as e:
            st.error(f"施設情報CSVファイルのエラー: {str(e)}")
        except Exception as e:
            st.error(f"施設情報CSVファイルの予期せぬエラー: {str(e)}")
    
    # 先月分のデータ（Excelファイル）のアップロード
    st.write("**3. 先月分のデータ（xlsx）をアップロードしてください。**")
    last_month_file = st.file_uploader("例: 【北九州市様】育児サークル等修正用データ（00月分）.xlsx", type=['xlsx'])
    if last_month_file:
        try:
            # Excelファイルの検証と読み込み
            import time
            overall_start = time.time()
            last_month_data = validate_excel_file(last_month_file)
            
            # データの整合性チェック（スラッグの一致確認）
            consistency_start = time.time()
            if 'circle_data' in locals() and circle_data is not None:
                check_data_consistency(circle_data, last_month_data)
            consistency_time = time.time() - consistency_start
            
            overall_time = time.time() - overall_start
            
            st.success("先月分のExcelファイルが正常に読み込まれました")
            
            # デバッグモード時に追加情報を表示
            if st.session_state.get('debug_mode', False):
                st.write(f"**📊 データ情報:** 行数: {len(last_month_data)}, 列数: {len(last_month_data.columns)}")
                st.write(f"**🔍 整合性チェック時間:** {consistency_time:.3f}秒")
                st.write(f"**⏱️ 全体処理時間:** {overall_time:.3f}秒")
            
            with st.expander("先月情報データを確認する"):
                st.dataframe(last_month_data, use_container_width=True)
        except ValueError as e:
            st.error(f"先月分のExcelファイルのエラー: {str(e)}")
        except Exception as e:
            st.error(f"先月分のExcelファイルの予期せぬエラー: {str(e)}")
    
    # 全てのデータが揃っているか確認
    all_data_ready = (
        'circle_data' in locals() and circle_data is not None and
        'facility_data' in locals() and facility_data is not None and
        'last_month_data' in locals() and last_month_data is not None
    )
    
    if all_data_ready:
        st.success("全てのファイルが正常に読み込まれました。処理を開始できます。")
        
        st.header('STEP2：データ処理の実行', divider='orange')
        
        # 自治体名の入力フィールドを追加（デフォルト値：北九州市様）
        municipality = st.text_input("自治体名", value="北九州市様", help="ダウンロードファイル名に使用される自治体名を入力してください")
        
        # 処理内容を事前に確認
        with st.expander("処理内容を確認する"):
            st.write("以下の処理が実行されます：")
            st.write("1. 0/1の値を持つ列の変換処理（○/空欄への変換）")
            st.write("2. 場所列の追加（施設情報との突合）")
            st.write("3. アカウント情報の追加（先月分データとの突合）")
        
        if st.button("処理開始"):
            try:
                # ファイル処理を実行
                output, proc_time, process_df, processed_circle_data = process_files(
                    circle_data,
                    facility_data=facility_data,
                    last_month_data=last_month_data
                )
                
                # デバッグモード時のみ処理時間と行数を表示
                if st.session_state.get('debug_mode', False):
                    st.info(f"処理時間: {proc_time:.3f}秒")
                    template_wb = load_workbook(TEMPLATE_FILE)
                    template_ws = template_wb.active
                    st.info(f"処理したデータ行数: {len(circle_data)-1}行")  # ヘッダーを除く
                    st.info(f"CSVファイルの列数: {len(circle_data.columns)}列")
                    st.info(f"テンプレートファイルの列数: {template_ws.max_column}列")
                
                st.success("ファイルの処理が完了しました！")
                
                st.header('STEP3：処理結果の確認', divider='orange')
                
                # 処理後のデータフレームを表示
                with st.expander("処理後のデータフレームを確認する"):
                    st.dataframe(processed_circle_data, use_container_width=True)
                
                st.header('STEP4：ダウンロード', divider='orange')
                
                # 現在の月を取得
                current_month = datetime.datetime.now().month
                
                # ファイル名を生成
                file_name = f"【{municipality}】育児サークル等修正用データ（{current_month}月分）.xlsx"
                
                # ダウンロードボタンを表示
                st.download_button(
                    label="処理済みファイルをダウンロード",
                    data=output,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except ValueError as e:
                st.error(str(e))
            except Exception as e:
                st.error(f"予期せぬエラーが発生しました: {str(e)}")

def validate_modification_status(main_data, original_data):
    """修正・削除新規ステータスの検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
        original_data (pd.DataFrame): 差分検出用データ
    
    Returns:
        pd.DataFrame: エラー情報を含むデータフレーム
    """
    errors = []
    valid_statuses = ['修正', '新規追加', '掲載順', '削除']
    
    for idx, row in main_data.iterrows():
        error_list = []
        status = normalize_value(row.get('修正・削除新規', ''))
        
        # ステータス値の検証
        if status != '' and status not in valid_statuses:
            column_pos = get_column_position_text(main_data, '修正・削除新規')
            error_list.append(f"{column_pos}修正・削除新規列に、次の値以外が入力されています。(修正・新規追加・掲載順・削除)")
        
        # 修正ステータスの検証
        if status == '修正':
            slug = str(row.get('スラッグ', '')).strip()
            if slug:
                # 差分検出用データから同じスラッグの行を取得
                original_row = original_data[original_data['スラッグ'] == slug]
                if not original_row.empty:
                    # 修正・削除新規列とアカウント関連列以外の列で差分をチェック
                    excluded_columns = ['修正・削除新規', 'ｱｶｳﾝﾄ発行有無', 'ｱｶｳﾝﾄ発行年月', 'アカウント発行の登録用メールアドレス']
                    check_columns = [col for col in main_data.columns if col not in excluded_columns]
                    has_difference = False
                    
                    for col in check_columns:
                        if col in original_row.columns:
                            main_value = normalize_value(row.get(col, ''))
                            original_value = normalize_value(original_row.iloc[0].get(col, ''))
                            
                            if main_value != original_value:
                                has_difference = True
                                break
                    
                    if not has_difference:
                        # アカウント関連のみの変更の場合はエラーとしない
                        if not is_only_account_related_change(row, original_data):
                            column_pos = get_column_position_text(main_data, '修正・削除新規')
                            error_list.append(f"{column_pos}修正にもかかわらず、値が変更されていません")
        
        # 新規追加ステータスの検証
        elif status == '新規追加':
            slug = normalize_value(row.get('スラッグ', ''))
            if slug != '':
                slug_column_pos = get_column_position_text(main_data, 'スラッグ')
                error_list.append(f"{slug_column_pos}新規追加にもかかわらずスラッグ列に値が入力されています")
            
            # HP掲載可列の検証
            hp_publish = normalize_value(row.get('HP掲載可', ''))
            if hp_publish != '○':
                modification_column_pos = get_column_position_text(main_data, '修正・削除新規')
                hp_column_pos = get_column_position_text(main_data, 'HP掲載可')
                error_list.append(f"{modification_column_pos}修正・削除新規列が「新規追加」ですが{hp_column_pos}HP掲載可列の値が「○」ではありません")
        
        # 掲載順ステータスの検証
        elif status == '掲載順':
            slug = str(row.get('スラッグ', '')).strip()
            if slug:
                original_row = original_data[original_data['スラッグ'] == slug]
                if not original_row.empty:
                    main_order = normalize_value(row.get('順番', ''))
                    original_order = normalize_value(original_row.iloc[0].get('順番', ''))
                    
                    if main_order == original_order:
                        order_column_pos = get_column_position_text(main_data, '順番')
                        error_list.append(f"{order_column_pos}「掲載順」ステータスが振られていますが、順番が変わっていません")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_empty_status(main_data, original_data):
    """空欄の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
        original_data (pd.DataFrame): 差分検出用データ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    
    for idx, row in main_data.iterrows():
        error_list = []
        status = normalize_value(row.get('修正・削除新規', ''))
        
        if status == '':  # 空欄または欠損値の場合
            slug = str(row.get('スラッグ', '')).strip()
            if slug:
                original_row = original_data[original_data['スラッグ'] == slug]
                if not original_row.empty:
                    # 修正・削除新規列とアカウント関連列以外の列で差分をチェック
                    excluded_columns = ['修正・削除新規', 'ｱｶｳﾝﾄ発行有無', 'ｱｶｳﾝﾄ発行年月', 'アカウント発行の登録用メールアドレス']
                    check_columns = [col for col in main_data.columns if col not in excluded_columns]
                    changed_columns = []
                    
                    for col in check_columns:
                        if col in original_row.columns:
                            main_value = normalize_value(row.get(col, ''))
                            original_value = normalize_value(original_row.iloc[0].get(col, ''))
                            
                            if main_value != original_value:
                                changed_columns.append(col)
                    
                    if changed_columns:
                        # 変更された列の位置情報を取得
                        changed_columns_with_pos = []
                        for col in changed_columns:
                            col_pos = get_column_position_text(main_data, col)
                            changed_columns_with_pos.append(f"{col_pos}{col}")
                        
                        error_list.append(f"修正と書かれていませんが、{','.join(changed_columns_with_pos)}の値が変更されています")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_machine_dependent_characters(main_data):
    """機種依存文字の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    target_columns = ['サークル名', '概要', '活動場所', '申込方法', '会費', '活動日_備考', 
                     '団体名（ふりがな）', '小学校区', '小学校区（ふりがな）', '代表者名', 
                     '代表者名（ふりがな）', '代表者住所', '記入者', '場所']
    
    # 機種依存文字のパターン（一部の例）
    machine_dependent_chars = ['①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩', 
                              '㍉', '㌔', '㌘', '㌧', '㌃', '㌍', '㌦', '㌢', '㌘', '㌧']
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        for col in target_columns:
            if col in main_data.columns:
                value = normalize_value(row.get(col, ''))
                
                if value:  # 空欄でない場合のみチェック
                    for char in machine_dependent_chars:
                        if char in value:
                            col_pos = get_column_position_text(main_data, col)
                            error_list.append(f"{col_pos}{col}列に機種依存文字が含まれています")
                            break
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_cell_line_breaks(main_data):
    """セル内改行の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    target_columns = ['サークル名', '活動種別', '活動場所', '申込方法', 'Eメールアドレス', '会費', 
                     'Webサイト', '活動日_備考', '団体名（ふりがな）', '幼稚園・保育園チェック', 
                     '小学校区', '小学校区（ふりがな）', '代表者名', '代表者名（ふりがな）', 
                     '代表者住所', '記入者', '場所']
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        for col in target_columns:
            if col in main_data.columns:
                value = normalize_value(row.get(col, ''))
                
                if value and ('\n' in value or '\r' in value):
                    col_pos = get_column_position_text(main_data, col)
                    error_list.append(f"{col_pos}{col}列にセル内改行が含まれています")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_prohibited_changes(main_data, original_data):
    """変更禁止列の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
        original_data (pd.DataFrame): 差分検出用データ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    prohibited_columns = ['スラッグ', 'ステータス', '参加者の条件(妊娠後半)', '参加者の条件(出産)', 
                         '参加者の条件(1歳後半)', '参加者の条件(2歳後半)', '申込方法備考', 
                         '活動日_営業時間ラベル', '代表者', '団体名']
    
    for idx, row in main_data.iterrows():
        error_list = []
        status = normalize_value(row.get('修正・削除新規', ''))
        
        # 新規追加の場合はスキップ
        if status == '新規追加':
            errors.append('')
            continue
        
        slug = str(row.get('スラッグ', '')).strip()
        if slug:
            original_row = original_data[original_data['スラッグ'] == slug]
            if not original_row.empty:
                changed_columns = []
                
                for col in prohibited_columns:
                    if col in main_data.columns and col in original_row.columns:
                        main_value = normalize_value(row.get(col, ''))
                        original_value = normalize_value(original_row.iloc[0].get(col, ''))
                        
                        if main_value != original_value:
                            changed_columns.append(col)
                
                if changed_columns:
                    # 変更された列の位置情報を取得
                    changed_columns_with_pos = []
                    for col in changed_columns:
                        col_pos = get_column_position_text(main_data, col)
                        changed_columns_with_pos.append(f"{col_pos}{col}")
                    
                    error_list.append(f"{','.join(changed_columns_with_pos)}の値が変更されています")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_consecutive_spaces(main_data):
    """連続した空白の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    target_columns = ['サークル名', '概要', '活動場所', '申込方法', '会費', '活動日_備考', 
                     '団体名（ふりがな）', '小学校区', '小学校区（ふりがな）', '代表者名', 
                     '代表者名（ふりがな）', '代表者住所', '記入者', '場所']
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        for col in target_columns:
            if col in main_data.columns:
                value = normalize_value(row.get(col, ''))
                
                if value and '   ' in value:  # 3つ以上の連続した空白
                    col_pos = get_column_position_text(main_data, col)
                    error_list.append(f"{col_pos}{col}列に連続した空白が含まれています")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_alphanumeric(main_data):
    """半角英数の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    # Webサイト列を除外（URL検証で別途処理）
    target_columns = ['申込先電話番号', '代表者郵便番号', '代表者電話番号', 
                     '代表者FAX', '代表者携帯番号', '順番']
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        for col in target_columns:
            if col in main_data.columns:
                value = normalize_value(row.get(col, ''))
                
                if value:  # 空欄でない場合のみチェック
                    # 半角英数字、各種ハイフン、ピリオド、スラッシュ、コロンのみ許可
                    if not re.match(r'^[a-zA-Z0-9\-‐–—−\.\/:]*$', value):
                        col_pos = get_column_position_text(main_data, col)
                        error_list.append(f"{col_pos}{col}列に半角英数字以外の文字が含まれています")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_email_addresses(main_data):
    """メールアドレスの検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    target_columns = ['Eメールアドレス', 'アカウント発行の登録用メールアドレス']
    
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        for col in target_columns:
            if col in main_data.columns:
                value = normalize_value(row.get(col, ''))
                
                if value and not email_pattern.match(value):
                    col_pos = get_column_position_text(main_data, col)
                    error_list.append(f"{col_pos}{col}列のメールアドレスが無効です")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_required_fields(main_data):
    """必須項目の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    required_columns = ['サークル名', 'スラッグ', 'ステータス', '活動種別']
    
    for idx, row in main_data.iterrows():
        error_list = []
        status = normalize_value(row.get('修正・削除新規', ''))
        
        # 新規追加の場合はスキップ
        if status == '新規追加':
            errors.append('')
            continue
        
        for col in required_columns:
            if col in main_data.columns:
                value = normalize_value(row.get(col, ''))
                
                if not value:
                    col_pos = get_column_position_text(main_data, col)
                    error_list.append(f"{col_pos}{col}列が空欄です")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_circle_or_cross(main_data):
    """マルバツの検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    target_columns = ['参加者の条件(妊娠)', '参加者の条件(妊娠後半)', '参加者の条件(0歳)', 
                     '参加者の条件(1歳)', '参加者の条件(1歳後半)', '参加者の条件(2歳)', 
                     '参加者の条件(2歳後半)', '参加者の条件(3歳)', '参加者の条件(4歳)', 
                     '参加者の条件(5歳)', '参加者の条件(6歳)', '参加者の条件(どなたでも)', 
                     '要会費', '冊子掲載可', 'HP掲載可', 'オープンデータ掲載可']
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        for col in target_columns:
            if col in main_data.columns:
                value = normalize_value(row.get(col, ''))
                
                if value and value not in ['○', '']:
                    col_pos = get_column_position_text(main_data, col)
                    error_list.append(f"{col_pos}{col}列に○または空欄以外の値が入力されています")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

async def validate_website_urls(main_data):
    """webサイトURL検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    try:
        from validate import is_url_alive
        import aiohttp
    except ImportError as e:
        # validate.pyが存在しない場合は検証をスキップ
        st.warning(f"WebサイトURL検証をスキップします: {str(e)}")
        return [''] * len(main_data)
    
    errors = []
    target_column = 'Webサイト'
    
    if target_column not in main_data.columns:
        return [''] * len(main_data)
    
    # 空でないURLのみを抽出
    urls_to_check = []
    for idx, row in main_data.iterrows():
        raw_value = row.get(target_column, '')
        # 空欄と欠損値を同じものとして扱う
        value = normalize_value(raw_value)
        
        if value:  # 空欄でない場合のみチェック
            # @で始まる場合は@を除去
            if value.startswith('@'):
                value = value[1:]
            urls_to_check.append((idx, value))
        else:
            errors.append('')
    
    if not urls_to_check:
        return errors
    
    # プログレスバーとステータステキストを表示（URL数が2以上の場合のみ）
    progress_bar = None
    status_text = None
    total_urls = len(urls_to_check)
    
    if total_urls >= 2:
        st.info(f"WebサイトURL検証を開始します（{total_urls}件のURLを検証）")
        progress_bar = st.progress(0)
        status_text = st.empty()
        status_text.text(f"WebサイトURL検証中: 0/{total_urls}")
    
    # 非同期でURL検証を実行
    try:
        async with aiohttp.ClientSession() as session:
            for current_index, (idx, url) in enumerate(urls_to_check):
                try:
                    # プログレスバーの更新
                    if progress_bar is not None:
                        progress = (current_index + 1) / total_urls
                        progress_bar.progress(progress)
                        status_text.text(f"WebサイトURL検証中: {current_index + 1}/{total_urls} - {url[:50]}{'...' if len(url) > 50 else ''}")
                    
                    _, error_msg = await is_url_alive(url, target_column, session)
                    if idx >= len(errors):
                        errors.extend([''] * (idx - len(errors) + 1))
                    
                    # エラーメッセージに列位置を追加
                    if error_msg:
                        col_pos = get_column_position_text(main_data, target_column)
                        # 既存のエラーメッセージから列名部分を除去して列位置を追加
                        if error_msg.startswith(f'{target_column}列で'):
                            error_msg = error_msg.replace(f'{target_column}列で', f'{col_pos}{target_column}列で')
                        elif error_msg.startswith(f'{target_column}列'):
                            error_msg = error_msg.replace(f'{target_column}列', f'{col_pos}{target_column}列')
                        else:
                            error_msg = f'{col_pos}{error_msg}'
                    
                    errors[idx] = error_msg
                except Exception as e:
                    if idx >= len(errors):
                        errors.extend([''] * (idx - len(errors) + 1))
                    col_pos = get_column_position_text(main_data, target_column)
                    errors[idx] = f"{col_pos}{target_column}列でURL検証エラー: {str(e)}"
    except Exception as e:
        # aiohttp関連のエラーの場合
        st.warning(f"WebサイトURL検証でエラーが発生しました: {str(e)}")
        # エラーが発生した場合は空のエラーリストを返す
        for idx, _ in urls_to_check:
            if idx >= len(errors):
                errors.extend([''] * (idx - len(errors) + 1))
    finally:
        # プログレスバーとステータステキストをクリア
        if progress_bar is not None:
            progress_bar.empty()
        if status_text is not None:
            status_text.empty()
    
    # 不足分を空文字で埋める
    while len(errors) < len(main_data):
        errors.append('')
    
    return errors

def validate_facility_location(main_data, facility_data):
    """活動場所の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
        facility_data (pd.DataFrame): 施設情報データ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    
    if facility_data is None or '施設名' not in facility_data.columns:
        # 施設情報がない場合はスキップ
        return [''] * len(main_data)
    
    facility_names = set(facility_data['施設名'].astype(str).str.strip())
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        if '活動場所' in main_data.columns:
            raw_value = row.get('活動場所', '')
            # 空欄と欠損値を同じものとして扱う
            if pd.isna(raw_value):
                value = ''
            else:
                value = str(raw_value).strip()
                if value == 'nan' or value == 'None' or value == '<NA>':
                    value = ''
            
            # 空欄でない場合のみチェック
            if value and value not in facility_names:
                col_pos = get_column_position_text(main_data, '活動場所')
                error_list.append(f"{col_pos}活動場所が施設情報に存在しません")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_status_column(main_data):
    """ステータス列の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    valid_statuses = ['publish', 'private', '']
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        if 'ステータス' in main_data.columns:
            raw_value = row.get('ステータス', '')
            # 空欄と欠損値を同じものとして扱う
            if pd.isna(raw_value):
                value = ''
            else:
                value = str(raw_value).strip()
                if value == 'nan' or value == 'None' or value == '<NA>':
                    value = ''
            
            if value not in valid_statuses:
                col_pos = get_column_position_text(main_data, 'ステータス')
                error_list.append(f"{col_pos}ステータス列に無効な値が入力されています")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_account_issue_date(main_data):
    """アカウント発行年月の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    
    # 和暦から西暦への変換関数（検証用）
    def convert_wareki_to_seireki_for_validation(wareki_str):
        if pd.isna(wareki_str):
            return None
        
        # 文字列に変換して正規化
        wareki_str = str(wareki_str).strip()
        if not wareki_str or wareki_str in ['nan', 'None', '<NA>']:
            return None
            
        try:
            # カンマまたはピリオドで分割
            separator = ',' if ',' in wareki_str else '.' if '.' in wareki_str else None
            if separator:
                parts = wareki_str.split(separator)
                if len(parts) == 2:
                    year_part = parts[0].strip()
                    month_part = int(parts[1].strip())
                    
                    # 月の範囲チェック
                    if not (1 <= month_part <= 12):
                        return False  # 無効な月
                    
                    if year_part.startswith('R'):
                        # 令和
                        reiwa_year = int(year_part[1:])
                        # 令和年の妥当性チェック（令和1年〜令和50年程度まで）
                        if not (1 <= reiwa_year <= 50):
                            return False  # 無効な令和年
                        seireki_year = 2018 + reiwa_year
                        return seireki_year * 100 + month_part
            return False  # 変換できない形式
        except:
            return False  # 変換エラー
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        if 'ｱｶｳﾝﾄ発行年月' in main_data.columns:
            value = normalize_value(row.get('ｱｶｳﾝﾄ発行年月', ''))
            
            # 空欄でない場合のみ検証
            if value:
                conversion_result = convert_wareki_to_seireki_for_validation(value)
                if conversion_result is False:
                    col_pos = get_column_position_text(main_data, 'ｱｶｳﾝﾄ発行年月')
                    error_list.append(f"{col_pos}ｱｶｳﾝﾄ発行年月列に変換できない値が入力されています")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_weekdays(main_data):
    """曜日の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    target_column = '活動日_営業曜日'
    
    if target_column not in main_data.columns:
        return [''] * len(main_data)
    
    valid_days = {'月', '火', '水', '木', '金', '土', '日', '祝'}
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        value = normalize_value(row.get(target_column, ''))
        
        # 空欄でない場合のみチェック
        if value:
            try:
                days = set(value.split(','))  # カンマで分割してセットに変換
                if not days.issubset(valid_days):
                    col_pos = get_column_position_text(main_data, target_column)
                    error_list.append(f"{col_pos}活動日_営業曜日列はカンマ区切りで入力してください")
            except AttributeError:
                col_pos = get_column_position_text(main_data, target_column)
                error_list.append(f"{col_pos}活動日_営業曜日列はカンマ区切りで入力してください")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def validate_business_hours(main_data):
    """時間の検証
    
    Args:
        main_data (pd.DataFrame): メインデータ
    
    Returns:
        list: エラーメッセージのリスト
    """
    errors = []
    start_column = '活動日_開始時間'
    end_column = '活動日_終了時間'
    
    # 列が存在しない場合は空のエラーリストを返す
    if start_column not in main_data.columns or end_column not in main_data.columns:
        return [''] * len(main_data)
    
    def is_valid_time_format(time_str):
        """時間形式が正しいかチェック"""
        if not time_str:
            return False
        try:
            # HH:MM または HH:MM:SS 形式をチェック
            if ':' not in time_str:
                return False
            
            parts = time_str.split(':')
            if len(parts) == 2:  # HH:MM
                hours, minutes = map(int, parts)
                return 0 <= hours <= 23 and 0 <= minutes <= 59
            elif len(parts) == 3:  # HH:MM:SS
                hours, minutes, seconds = map(int, parts)
                return 0 <= hours <= 23 and 0 <= minutes <= 59 and 0 <= seconds <= 59
            else:
                return False
        except (ValueError, TypeError):
            return False
    
    def time_to_minutes(time_str):
        """時間文字列を分に変換（比較用）"""
        try:
            parts = time_str.split(':')
            hours = int(parts[0])
            minutes = int(parts[1])
            return hours * 60 + minutes
        except (ValueError, IndexError):
            return None
    
    for idx, row in main_data.iterrows():
        error_list = []
        
        start_value = normalize_value(row.get(start_column, ''))
        end_value = normalize_value(row.get(end_column, ''))
        
        # 両方空欄の場合は検証しない
        if not start_value and not end_value:
            errors.append('')
            continue
        
        # 開始時間の形式チェック
        start_valid = is_valid_time_format(start_value) if start_value else True
        end_valid = is_valid_time_format(end_value) if end_value else True
        
        if start_value and not start_valid:
            if end_value and not end_valid:
                start_col_pos = get_column_position_text(main_data, start_column)
                end_col_pos = get_column_position_text(main_data, end_column)
                error_list.append(f"{start_col_pos}開始+{end_col_pos}終了時間の形式が違います")
            else:
                start_col_pos = get_column_position_text(main_data, start_column)
                error_list.append(f"{start_col_pos}開始時間の形式が違います")
        elif end_value and not end_valid:
            end_col_pos = get_column_position_text(main_data, end_column)
            error_list.append(f"{end_col_pos}終了時間の形式が違います")
        elif start_value and end_value and start_valid and end_valid:
            # 開始時間と終了時間の論理チェック
            start_minutes = time_to_minutes(start_value)
            end_minutes = time_to_minutes(end_value)
            
            if start_minutes is not None and end_minutes is not None:
                if start_minutes >= end_minutes:
                    start_col_pos = get_column_position_text(main_data, start_column)
                    end_col_pos = get_column_position_text(main_data, end_column)
                    error_list.append(f"{start_col_pos}開始時間と{end_col_pos}終了時間が同じまたは逆転しています")
        
        errors.append(', '.join(error_list) if error_list else '')
    
    return errors

def perform_data_validation(main_data, original_data, facility_data=None, validation_options=None):
    """データ検証の実行
    
    Args:
        main_data (pd.DataFrame): メインデータ
        original_data (pd.DataFrame): 差分検出用データ
        facility_data (pd.DataFrame, optional): 施設情報データ
        validation_options (dict, optional): 実行する検証項目の選択
    
    Returns:
        pd.DataFrame: エラー情報を含むメインデータ
    """
    # エラー列を初期化
    main_data_with_errors = main_data.copy()
    
    # デフォルトでは全ての検証を実行
    if validation_options is None:
        validation_options = {
            'modification_status': True,
            'empty_status': True,
            'machine_dependent': True,
            'cell_breaks': True,
            'prohibited_changes': True,
            'consecutive_spaces': True,
            'alphanumeric': True,
            'email': True,
            'required_fields': True,
            'circle_cross': True,
            'facility_location': True,
            'status_column': True,
            'website_urls': True,
            'account_issue_date': True,
            'weekdays': True,
            'business_hours': True
        }
    
    # 各検証を実行
    validation_functions = [
        ('modification_status', '修正・削除新規ステータス', lambda: validate_modification_status(main_data, original_data)),
        ('empty_status', '空欄ステータス', lambda: validate_empty_status(main_data, original_data)),
        ('machine_dependent', '機種依存文字', lambda: validate_machine_dependent_characters(main_data)),
        ('cell_breaks', 'セル内改行', lambda: validate_cell_line_breaks(main_data)),
        ('prohibited_changes', '変更禁止列', lambda: validate_prohibited_changes(main_data, original_data)),
        ('consecutive_spaces', '連続した空白', lambda: validate_consecutive_spaces(main_data)),
        ('alphanumeric', '半角英数', lambda: validate_alphanumeric(main_data)),
        ('email', 'メールアドレス', lambda: validate_email_addresses(main_data)),
        ('required_fields', '必須項目', lambda: validate_required_fields(main_data)),
        ('circle_cross', 'マルバツ', lambda: validate_circle_or_cross(main_data)),
        ('facility_location', '活動場所', lambda: validate_facility_location(main_data, facility_data)),
        ('status_column', 'ステータス', lambda: validate_status_column(main_data)),
        ('account_issue_date', 'アカウント発行年月', lambda: validate_account_issue_date(main_data)),
        ('weekdays', '曜日', lambda: validate_weekdays(main_data)),
        ('business_hours', '時間', lambda: validate_business_hours(main_data))
    ]
    
    # 非同期検証（webサイトURL検証）
    async_validation_functions = [
        ('website_urls', 'webサイトURL', lambda: validate_website_urls(main_data))
    ]
    
    all_errors = []
    executed_validations = []
    
    # 同期検証を実行
    for validation_key, validation_name, validation_func in validation_functions:
        if validation_options.get(validation_key, False):
            try:
                errors = validation_func()
                all_errors.append(errors)
                executed_validations.append(validation_name)
            except Exception as e:
                st.error(f"{validation_name}の検証中にエラーが発生しました: {str(e)}")
                all_errors.append([''] * len(main_data))
                executed_validations.append(f"{validation_name}（エラー）")
    
    # 非同期検証を実行
    import asyncio
    for validation_key, validation_name, validation_func in async_validation_functions:
        if validation_options.get(validation_key, False):
            try:
                errors = asyncio.run(validation_func())
                all_errors.append(errors)
                executed_validations.append(validation_name)
            except Exception as e:
                st.error(f"{validation_name}の検証中にエラーが発生しました: {str(e)}")
                all_errors.append([''] * len(main_data))
                executed_validations.append(f"{validation_name}（エラー）")
    
    # 実行された検証項目を表示
    if executed_validations:
        st.info(f"実行された検証項目: {', '.join(executed_validations)}")
    else:
        st.warning("検証項目が選択されていません。")
        main_data_with_errors['エラー'] = [''] * len(main_data)
        return main_data_with_errors
    
    # 全てのエラーを統合
    combined_errors = []
    for i in range(len(main_data)):
        row_errors = []
        for error_list in all_errors:
            if i < len(error_list) and error_list[i]:
                row_errors.append(error_list[i])
        combined_errors.append(', '.join(row_errors))
    
    main_data_with_errors['エラー'] = combined_errors
    
    return main_data_with_errors

def validate_import_excel_file(excel_file, skip_rows_count=2):
    """インポート用Excelファイルの検証と読み込みを行う
    
    Args:
        excel_file: アップロードされたExcelファイル
        skip_rows_count: スキップする行数
    
    Returns:
        tuple: (メインデータ, 差分検出用データ)
    """
    import time
    import streamlit as st
    
    # 開始時間を記録
    start_time = time.time()
    timing_info = []
    
    try:
        # Excelファイルを読み込んでシート情報を取得
        sheet_info_start = time.time()
        wb = pd.ExcelFile(excel_file)
        sheet_names = wb.sheet_names
        sheet_info_time = time.time() - sheet_info_start
        timing_info.append(f"シート情報取得: {sheet_info_time:.3f}秒")
        
        # シート数の検証
        validation_start = time.time()
        if len(sheet_names) > 2:
            raise ValueError("シート数が2より多いため、どのシートをメインデータにするかが特定できません")
        
        if len(sheet_names) < 2:
            raise ValueError("originalシートと別のシートが必要ですが、シート数が不足しています")
        
        # originalシートと別シートを特定
        original_sheet = None
        main_sheet = None
        
        for sheet_name in sheet_names:
            if sheet_name.lower() == 'original':
                original_sheet = sheet_name
            else:
                main_sheet = sheet_name
        
        if original_sheet is None:
            raise ValueError("'original'という名前のシートが見つかりません")
        
        validation_time = time.time() - validation_start
        timing_info.append(f"シート検証: {validation_time:.3f}秒")
        
        # メインデータを読み込み（指定された行数をスキップ）
        main_data_start = time.time()
        main_data = pd.read_excel(excel_file, sheet_name=main_sheet, skiprows=list(range(1, skip_rows_count + 1)))
        main_data_time = time.time() - main_data_start
        timing_info.append(f"メインデータ読み込み: {main_data_time:.3f}秒")
        
        # 差分検出用データを読み込み（指定された行数をスキップ）
        original_data_start = time.time()
        original_data = pd.read_excel(excel_file, sheet_name=original_sheet, skiprows=list(range(1, skip_rows_count + 1)))
        original_data_time = time.time() - original_data_start
        timing_info.append(f"差分検出用データ読み込み: {original_data_time:.3f}秒")
        
        # 基本的な検証
        basic_validation_start = time.time()
        if main_data.empty:
            raise ValueError("メインデータが空です")
        if original_data.empty:
            raise ValueError("差分検出用データが空です")
            
        if len(main_data.columns) == 0:
            raise ValueError("メインデータに列が存在しません")
        if len(original_data.columns) == 0:
            raise ValueError("差分検出用データに列が存在しません")
        
        basic_validation_time = time.time() - basic_validation_start
        timing_info.append(f"基本検証: {basic_validation_time:.3f}秒")
        
        # 削除された行の検証（B列：スラッグで比較）
        deletion_check_start = time.time()
        if len(main_data.columns) >= 2 and len(original_data.columns) >= 2:
            # B列（インデックス1）がスラッグ列
            main_slugs = set(main_data.iloc[:, 1].dropna().astype(str))
            original_slugs = set(original_data.iloc[:, 1].dropna().astype(str))
            
            # 差分検出用データに存在するが、メインデータに存在しないスラッグを検出
            deleted_slugs = original_slugs - main_slugs
            
            if deleted_slugs:
                # 削除されたスラッグに対応する行データを取得
                deleted_rows = original_data[original_data.iloc[:, 1].astype(str).isin(deleted_slugs)]
                
                st.warning(f"⚠️ 以下の{len(deleted_slugs)}件のデータがメインデータから削除されています：")
                
                # 削除されたデータを表形式で表示
                if not deleted_rows.empty:
                    with st.expander(f"削除されたデータの詳細 ({len(deleted_rows)}件)"):
                        st.dataframe(deleted_rows, use_container_width=True)
        
        deletion_check_time = time.time() - deletion_check_start
        timing_info.append(f"削除行検証: {deletion_check_time:.3f}秒")
        
        total_time = time.time() - start_time
        timing_info.append(f"全体処理時間: {total_time:.3f}秒")
        
        # デバッグモード時に処理時間を表示
        if st.session_state.get('debug_mode', False):
            st.write("**⏱️ インポート用Excelファイル読み込み処理時間:**")
            for timing in timing_info:
                st.write(f"  - {timing}")
        
        return main_data, original_data
        
    except Exception as e:
        raise ValueError(f"Excelファイルの読み込み中にエラーが発生しました: {str(e)}")

def show_import_data_page():
    """インポートデータ作成ページの表示"""
    # st.header("インポートデータ作成", divider='orange')
    
    if st.session_state.debug_mode:
        st.write("デバッグモードが有効です")
    
    st.header('STEP1：ファイルのアップロード', divider='orange')
    
    # 修正済みExcelファイルのアップロード
    st.write("**1. 修正済みExcelファイル（xlsx）をアップロードしてください。**")
    excel_file = st.file_uploader("例: 【北九州市様】育児サークル等修正用データ（1月分）_修正済み.xlsx", type=['xlsx'], key="import_excel")
    
    # スキップする行数の指定
    skip_rows = st.number_input("スキップする行数", min_value=0, max_value=10, value=2, 
                               help="ヘッダー以外の上から何行をスキップするかを指定してください")
    
    main_data = None
    original_data = None
    
    if excel_file:
        try:
            # ファイルが変更された場合のみセッション状態をリセット
            import time
            session_reset_start = time.time()
            if check_file_changed(excel_file, 'excel'):
                reset_import_session_state()
            session_reset_time = time.time() - session_reset_start
            
            # Excelファイルの検証と読み込み
            overall_start = time.time()
            main_data, original_data = validate_import_excel_file(excel_file, skip_rows)
            overall_time = time.time() - overall_start
            
            st.success("Excelファイルが正常に読み込まれました")
            
            # デバッグモード時に追加情報を表示
            if st.session_state.get('debug_mode', False):
                st.write(f"**🔄 セッション状態リセット時間:** {session_reset_time:.3f}秒")
                st.write(f"**⏱️ 全体処理時間:** {overall_time:.3f}秒")
            
            col1, col2 = st.columns(2)
            with col1:
                with st.expander("メインデータを確認する"):
                    st.dataframe(main_data, use_container_width=True)
                    st.info(f"行数: {len(main_data)}, 列数: {len(main_data.columns)}")
            
            with col2:
                with st.expander("差分検出用データを確認する"):
                    st.dataframe(original_data, use_container_width=True)
                    st.info(f"行数: {len(original_data)}, 列数: {len(original_data.columns)}")
                    
        except ValueError as e:
            st.error(f"Excelファイルのエラー: {str(e)}")
        except Exception as e:
            st.error(f"Excelファイルの予期せぬエラー: {str(e)}")
    
    # 施設情報CSVファイルのアップロード（データ検証用）
    st.write("**2. 施設情報（csv）をアップロードしてください。**")
    facility_csv_file = st.file_uploader("例: kitakyushu-city_facility_00000000000.csv", type=['csv'], key="import_facility")
    facility_data = None
    
    if facility_csv_file:
        try:
            # ファイルが変更された場合のみセッション状態をリセット
            import time
            session_reset_start = time.time()
            if check_file_changed(facility_csv_file, 'facility'):
                reset_import_session_state()
            session_reset_time = time.time() - session_reset_start
            
            # 施設情報CSVファイルの検証と読み込み（専用の検証関数を使用）
            overall_start = time.time()
            facility_data, facility_encoding, facility_debug_info = validate_facility_csv_file(facility_csv_file)
            overall_time = time.time() - overall_start
            
            st.success(f"施設情報CSVファイルが正常に読み込まれました（エンコーディング: {facility_encoding}）")
            
            # デバッグモード時に追加情報を表示
            if st.session_state.get('debug_mode', False):
                st.write(f"**🔄 セッション状態リセット時間:** {session_reset_time:.3f}秒")
                st.write(f"**⏱️ 全体処理時間:** {overall_time:.3f}秒")
            
            with st.expander("施設情報データを確認する"):
                st.dataframe(facility_data, use_container_width=True)
                
            # デバッグモード時に詳細情報を表示
            if st.session_state.get('debug_mode', False):
                with st.expander("🔍 施設情報CSV読み込み詳細"):
                    st.write("**エンコーディング検出ログ:**")
                    for info in facility_debug_info:
                        st.text(info)
        except ValueError as e:
            st.error(f"施設情報CSVファイルのエラー: {str(e)}")
        except Exception as e:
            st.error(f"施設情報CSVファイルの予期せぬエラー: {str(e)}")
    
    # ユーザーデータCSVファイルのアップロード（インポートデータ作成用）
    st.write("**3. ユーザーデータ（csv）をアップロードしてください。**")
    user_csv_file = st.file_uploader("例: kitakyushu-city_user_00000000000.csv", type=['csv'], key="import_user")
    user_data = None
    
    if user_csv_file:
        try:
            # ファイルが変更された場合のみセッション状態をリセット
            import time
            session_reset_start = time.time()
            if check_file_changed(user_csv_file, 'user'):
                reset_import_session_state()
            session_reset_time = time.time() - session_reset_start
            
            # ユーザーデータCSVファイルの検証と読み込み
            overall_start = time.time()
            user_data, user_encoding, user_debug_info = validate_csv_file(user_csv_file)
            overall_time = time.time() - overall_start
            
            st.success(f"ユーザーデータCSVファイルが正常に読み込まれました（エンコーディング: {user_encoding}）")
            
            # デバッグモード時に追加情報を表示
            if st.session_state.get('debug_mode', False):
                st.write(f"**🔄 セッション状態リセット時間:** {session_reset_time:.3f}秒")
                st.write(f"**⏱️ 全体処理時間:** {overall_time:.3f}秒")
            
            with st.expander("ユーザーデータを確認する"):
                st.dataframe(user_data, use_container_width=True)
                
            # デバッグモード時に詳細情報を表示
            if st.session_state.get('debug_mode', False):
                with st.expander("🔍 ユーザーデータCSV読み込み詳細"):
                    st.write("**エンコーディング検出ログ:**")
                    for info in user_debug_info:
                        st.text(info)
        except ValueError as e:
            st.error(f"ユーザーデータCSVファイルのエラー: {str(e)}")
        except Exception as e:
            st.error(f"ユーザーデータCSVファイルの予期せぬエラー: {str(e)}")
    
    # 全てのデータが揃っているか確認
    all_data_ready = (
        main_data is not None and
        original_data is not None and
        facility_data is not None and
        user_data is not None
    )
    
    if all_data_ready:
        st.success("全てのファイルが正常に読み込まれました。データ検証を開始できます。")
        
        st.header('STEP2：データ検証の実行', divider='orange')
        
        # 自治体名の入力フィールド
        municipality = st.text_input("自治体名", value="北九州市", help="インポートファイル名に使用される自治体名を入力してください", key="import_municipality")
        
        # 検証項目の選択
        st.write("### 実施する検証項目を選択してください")
        
        # 検証項目の定義（キー: (表示名, デフォルト値, ヘルプテキスト)）
        validation_items = {
            'modification_status': ('修正・削除新規ステータス', True, '修正・削除新規列の値が正しく設定されているかを検証します。修正時の変更有無、新規追加時のスラッグ空欄、掲載順変更の妥当性をチェックします。'),
            'empty_status': ('空欄ステータス', True, '修正・削除新規列が空欄の場合に、実際にデータが変更されていないかを検証します。'),
            'machine_dependent': ('機種依存文字', True, 'サークル名、概要、活動場所などの文字列項目に機種依存文字（①②③など）が含まれていないかをチェックします。'),
            'cell_breaks': ('セル内改行', True, 'セル内に改行文字（\\n、\\r）が含まれていないかを検証します。データの表示崩れを防ぎます。'),
            'prohibited_changes': ('変更禁止列', True, 'スラッグ、ステータス、参加者の条件（後半）、申込方法備考などの変更禁止列が変更されていないかをチェックします。'),
            'consecutive_spaces': ('連続した空白', True, '文字列項目に3つ以上の連続した空白が含まれていないかを検証します。'),
            'alphanumeric': ('半角英数', True, '電話番号、郵便番号、順番などの項目が半角英数字で入力されているかを検証します。'),
            'email': ('メールアドレス', True, 'メールアドレス項目が正しい形式で入力されているかを検証します。'),
            'required_fields': ('必須項目', True, 'サークル名、スラッグ、ステータス、活動種別などの必須項目が入力されているかをチェックします。'),
            'circle_cross': ('マルバツ', True, '参加者の条件、要会費、掲載可能性などの項目が○または空欄で入力されているかを検証します。'),
            'facility_location': ('活動場所', True, '活動場所に入力された施設名が施設情報データに存在するかを検証します。'),
            'status_column': ('ステータス', True, 'ステータス列の値がpublish、private、または空欄のいずれかであるかを検証します。'),
            'website_urls': ('webサイトURL', True, 'WebサイトURLが有効で、実際にアクセス可能かを検証します。（時間がかかる場合があります）'),
            'weekdays': ('曜日', True, '活動日_営業曜日列がカンマ区切りの正しい曜日形式（月,火,水など）で入力されているかを検証します。'),
            'business_hours': ('時間', True, '活動日_開始時間と活動日_終了時間がHH:MM形式で入力され、開始時間が終了時間より前であることを検証します。')
        }
        
        # チェックボックスを3列に均等配置
        validation_states = {}
        items_list = list(validation_items.items())
        columns = st.columns(3)
        
        # 項目を3列に分散配置
        for i, (key, (display_name, default_value, help_text)) in enumerate(items_list):
            col_index = i % 3  # 順番に列を循環
            with columns[col_index]:
                # セッション状態に値があればそれを使用、なければデフォルト値を使用
                checkbox_value = st.session_state.get(f"check_{key}", default_value)
                validation_states[key] = st.checkbox(
                    display_name,
                    value=checkbox_value,
                    help=help_text,
                    key=f"check_{key}"
                )
        
        # 全選択・全解除ボタン用のコールバック関数
        def select_all_callback():
            for key in validation_items.keys():
                st.session_state[f"check_{key}"] = True
        
        def deselect_all_callback():
            for key in validation_items.keys():
                st.session_state[f"check_{key}"] = False
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 4])
        with col_btn1:
            st.button("全選択", key="select_all", on_click=select_all_callback)
        
        with col_btn2:
            st.button("全解除", key="deselect_all", on_click=deselect_all_callback)
        
        if st.button("データ検証開始", key="start_validation"):
            try:
                with st.spinner("データ検証を実行中..."):
                    # 選択された検証項目を取得
                    validation_options = {key: validation_states[key] for key in validation_items.keys()}
                    
                    # データ検証を実行
                    validated_data = perform_data_validation(main_data, original_data, facility_data, validation_options)
                    
                    # セッション状態に保存
                    st.session_state.validated_data = validated_data
                    st.session_state.validation_completed = True
                    log_session_state_change("validation_completed", {
                        'data_rows': len(validated_data),
                        'selected_validations': list(validation_options.keys())
                    })
                    
                    # エラーがある行の数を計算
                    error_rows = validated_data[validated_data['エラー'] != '']
                    error_count = len(error_rows)
                    
            except Exception as e:
                st.error(f"データ検証中にエラーが発生しました: {str(e)}")
                if st.session_state.get('debug_mode', False):
                    st.exception(e)
        
        # セッション状態に基づいて検証結果を表示
        if st.session_state.validation_completed and st.session_state.validated_data is not None:
            validated_data = st.session_state.validated_data
            
            # エラーがある行の数を計算
            error_rows = validated_data[validated_data['エラー'] != '']
            error_count = len(error_rows)
            
            if error_count > 0:
                st.error(f"データ検証が完了しました。{error_count}件のエラーが見つかりました。")
                
                # エラー詳細の表示
                with st.expander(f"エラー詳細を確認する ({error_count}件)"):
                    st.dataframe(error_rows[['サークル名', 'スラッグ', 'エラー']], use_container_width=True)
                
                # 全データ（エラー列付き）の表示
                with st.expander("検証結果を確認する（全データ）"):
                    st.dataframe(validated_data, use_container_width=True)
                
                # エラー付きデータのダウンロード
                current_date = datetime.datetime.now().strftime("%Y%m%d")
                error_file_name = f"{municipality}_データ検証結果_{current_date}.xlsx"
                
                # Excelファイルとして出力
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    validated_data.to_excel(writer, sheet_name='検証結果', index=False)
                    error_rows.to_excel(writer, sheet_name='エラー一覧', index=False)
                
                output.seek(0)
                st.download_button(
                    label="検証結果をダウンロード",
                    data=output,
                    file_name=error_file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            else:
                st.success("データ検証が完了しました。エラーは見つかりませんでした。")
                
                # バルーンは一度だけ表示
                if not st.session_state.get('balloons_shown', False):
                    st.balloons()
                    st.session_state.balloons_shown = True
                
                # 全データの表示
                with st.expander("検証結果を確認する"):
                    st.dataframe(validated_data, use_container_width=True)
                
                # データ整形とインポートファイル作成
                st.header('STEP3：インポートデータ作成', divider='orange')
                
                # インポートデータ作成用のコールバック関数
                def create_import_data_callback():
                    try:
                        log_session_state_change("import_data_creation_started", {
                            'municipality': municipality
                        })
                        
                        # データを整形
                        formatted_data = format_for_import(main_data, original_data)
                        log_session_state_change("data_formatted", {
                            'formatted_rows': len(formatted_data)
                        })
                        
                        # インポートファイルを作成
                        import_files = create_import_files(formatted_data, original_data, user_data, municipality, main_data)
                        log_session_state_change("import_files_created", {
                            'file_count': len(import_files) if import_files else 0,
                            'filenames': list(import_files.keys()) if import_files else []
                        })
                        
                        # セッション状態に保存
                        st.session_state.import_files = import_files
                        st.session_state.formatted_data = formatted_data
                        st.session_state.import_data_created = True
                        log_session_state_change("import_data_creation_completed", {
                            'success': True
                        })
                        
                    except Exception as e:
                        log_session_state_change("import_data_creation_error", {
                            'error': str(e)
                        })
                        st.error(f"インポートデータ作成中にエラーが発生しました: {str(e)}")
                        if st.session_state.get('debug_mode', False):
                            st.exception(e)
                
                # インポートデータ作成ボタン
                st.button("インポートデータ作成開始", key="start_import_creation", on_click=create_import_data_callback)
                
                # インポートデータが作成済みの場合、結果を表示
                if st.session_state.import_data_created and st.session_state.import_files is not None:
                    import_files = st.session_state.import_files
                    formatted_data = st.session_state.formatted_data
                    
                    if import_files:
                        st.success(f"{len(import_files)}個のインポートファイルが作成されました。")
                        
                        # アカウント発行年月の警告メッセージがある場合は表示
                        if 'account_date_warning' in st.session_state:
                            st.warning(st.session_state.account_date_warning)
                            # 警告を表示したらセッション状態から削除（重複表示を防ぐ）
                            del st.session_state.account_date_warning
                        
                        # ユーザー作成の警告メッセージがある場合は表示
                        if 'user_creation_warning' in st.session_state:
                            st.warning(st.session_state.user_creation_warning)
                            # 警告を表示したらセッション状態から削除（重複表示を防ぐ）
                            del st.session_state.user_creation_warning
                        
                        # ユーザー修正の警告メッセージがある場合は表示
                        if 'user_modification_warning' in st.session_state:
                            st.warning(st.session_state.user_modification_warning)
                            # 警告を表示したらセッション状態から削除（重複表示を防ぐ）
                            del st.session_state.user_modification_warning
                        
                        # ユーザー情報の統合表示（修正と新規追加の両方）
                        if 'user_comprehensive_details' in st.session_state:
                            comprehensive_df = pd.DataFrame(st.session_state.user_comprehensive_details)
                            
                            # 処理種別でグループ化
                            modification_data = comprehensive_df[comprehensive_df['処理種別'] == '修正']
                            new_addition_data = comprehensive_df[comprehensive_df['処理種別'] == '新規追加']
                            total_count = len(comprehensive_df)
                            
                            with st.expander(f"👤 ユーザー情報の処理内容を確認する ({total_count}件)"):
                                if not modification_data.empty:
                                    st.markdown("#### 🔄 修正されるユーザー")
                                    st.dataframe(modification_data, use_container_width=True, hide_index=True)
                                    st.caption(f"💡 修正対象: {len(modification_data)}件のユーザー")
                                
                                if not new_addition_data.empty:
                                    st.markdown("#### ➕ 新規追加されるユーザー")
                                    st.dataframe(new_addition_data, use_container_width=True, hide_index=True)
                                    st.caption(f"💡 新規追加: {len(new_addition_data)}件のユーザー")
                                
                                # 総計表示
                                st.caption(f"📊 **合計: {total_count}件のユーザー処理** (修正: {len(modification_data)}件, 新規追加: {len(new_addition_data)}件)")
                                st.caption("⚠️ 上記のユーザー情報が処理されます。内容を確認してからダウンロードしてください。")
                            
                            # 表示したらセッション状態から削除（重複表示を防ぐ）
                            del st.session_state.user_comprehensive_details
                            
                            # 古い形式の修正詳細情報も削除（統合表示に移行したため）
                            if 'user_modification_details' in st.session_state:
                                del st.session_state.user_modification_details
                        
                        # すでに発行済みユーザー情報の表示
                        if 'already_issued_users' in st.session_state:
                            already_issued_df = pd.DataFrame(st.session_state.already_issued_users)
                            
                            with st.expander(f"ℹ️ 以下のユーザーはすでに登録がありました（インポートデータ作成不要） ({len(already_issued_df)}件)"):
                                st.dataframe(already_issued_df, use_container_width=True, hide_index=True)
                                st.caption(f"💡 すでに発行済み: {len(already_issued_df)}件のユーザー")
                                st.caption("ℹ️ これらのユーザーは既に発行されており、名前とメールアドレスに変更がないため、インポートデータの作成対象外です。")
                            
                            # 表示したらセッション状態から削除（重複表示を防ぐ）
                            del st.session_state.already_issued_users
                        
                        # 削除対象データの表示
                        deletion_data = formatted_data[formatted_data['修正・削除新規'] == '削除']
                        if not deletion_data.empty:
                            st.warning("### 🗑️ 削除対象データ")
                            st.write("以下のデータについてはインポートで消えないため、管理画面から**ボミ箱ポイ**を忘れずに")
                            
                            # 削除対象データの表示（重要な列のみ）
                            display_columns = ['サークル名', 'スラッグ', 'ステータス', '修正・削除新規']
                            available_columns = [col for col in display_columns if col in deletion_data.columns]
                            
                            st.dataframe(
                                deletion_data[available_columns], 
                                use_container_width=True, 
                                hide_index=True
                            )
                            st.caption(f"💡 削除対象: {len(deletion_data)}件のサークルデータ")
                            st.caption("⚠️ これらのデータはインポート後にステータスが「private」になりますが、完全に削除されるわけではありません。")
                            st.caption("📋 管理画面から手動でゴミ箱に移動する作業が必要です。")
                        
                        # インポート用CSVダウンロードセクションの見出し
                        st.header('STEP4：ファイルダウンロード', divider='orange')
                        st.write("作成されたインポート用CSVファイルをダウンロードしてください。")
                        
                        # 各ファイルのダウンロードボタンを表示
                        for filename, data in import_files.items():
                            # 修正CSVファイルの場合は特別な処理
                            if isinstance(data, dict) and 'display_data' in data and 'download_data' in data:
                                # 修正CSVファイルの場合
                                display_data = data['display_data']  # 表示用（修正対象列含む）
                                download_data = data['download_data']  # ダウンロード用（修正対象列除外）
                                
                                # ファイル内容のプレビュー（表示用データを使用）
                                with st.expander(f"📋 {filename} の内容を確認"):
                                    st.dataframe(display_data, use_container_width=True)
                                    st.info(f"行数: {len(display_data)}, 列数: {len(display_data.columns)}")
                                    st.caption("💡 「修正対象列」は内容確認用の列で、ダウンロードファイルには含まれません。")
                                
                                # CSVファイルとして出力（ダウンロード用データを使用）
                                csv_output = io.StringIO()
                                download_data.to_csv(csv_output, index=False, encoding='utf-8-sig')
                                csv_data = csv_output.getvalue().encode('utf-8-sig')
                                
                                st.download_button(
                                    label=f"{filename}をダウンロード",
                                    data=csv_data,
                                    file_name=filename,
                                    mime="text/csv",
                                    key=f"download_{filename}"
                                )
                            else:
                                # 通常のファイルの場合
                                # ファイル内容のプレビュー
                                with st.expander(f"📋 {filename} の内容を確認"):
                                    st.dataframe(data, use_container_width=True)
                                    st.info(f"行数: {len(data)}, 列数: {len(data.columns)}")
                                
                                # CSVファイルとして出力
                                csv_output = io.StringIO()
                                data.to_csv(csv_output, index=False, encoding='utf-8-sig')
                                csv_data = csv_output.getvalue().encode('utf-8-sig')
                                
                                st.download_button(
                                    label=f"{filename}をダウンロード",
                                    data=csv_data,
                                    file_name=filename,
                                    mime="text/csv",
                                    key=f"download_{filename}"
                                )

                    else:
                        st.warning("作成対象のインポートデータがありませんでした。")
    


def format_for_import(main_data, original_data):
    """インポートデータ用に整形
    
    Args:
        main_data (pd.DataFrame): メインデータ
        original_data (pd.DataFrame): 差分検出用データ
    
    Returns:
        pd.DataFrame: 整形後のデータ
    """
    formatted_data = main_data.copy()
    
    # 数字へ置換
    binary_columns = ['参加者の条件(妊娠)', '参加者の条件(0歳)', '参加者の条件(1歳)', 
                     '参加者の条件(1歳後半)', '参加者の条件(2歳)', '参加者の条件(2歳後半)', 
                     '参加者の条件(3歳)', '参加者の条件(4歳)', '参加者の条件(5歳)', 
                     '参加者の条件(6歳)', '参加者の条件(どなたでも)', '要会費', 
                     '冊子掲載可', 'HP掲載可', 'オープンデータ掲載可']
    
    for col in binary_columns:
        if col in formatted_data.columns:
            # 列を文字列型に変換（警告を回避）
            formatted_data[col] = formatted_data[col].astype(str)
            
            for idx, raw_value in formatted_data[col].items():
                # normalize_value関数を使用して統一的に処理
                value = normalize_value(raw_value)
                
                # 値の変換
                if value == '' or value == '0':
                    formatted_data.at[idx, col] = '0'
                elif value == '○' or value == '1':
                    formatted_data.at[idx, col] = '1'
                else:
                    # それ以外の値が入っていた場合はエラー扱い
                    st.error(f"行{idx+1}の{col}列に無効な値が含まれています: {value}")
                    formatted_data.at[idx, col] = '0'  # デフォルト値として0を設定
    
    # 参加者の条件(妊娠後半)列に参加者の条件(妊娠)列の値をコピー
    if '参加者の条件(妊娠)' in formatted_data.columns and '参加者の条件(妊娠後半)' in formatted_data.columns:
        formatted_data['参加者の条件(妊娠後半)'] = formatted_data['参加者の条件(妊娠)']
    
    # 入力禁止列の値の削除
    prohibited_columns = ['申込方法備考', '活動日_営業時間ラベル', '活動日_営業曜日ラベル']
    for col in prohibited_columns:
        if col in formatted_data.columns:
            formatted_data[col] = ''
    
    # 参加者の条件(出産)は一律「0」で埋める（入力禁止列だが、インポートデータでは「0」が必要）
    if '参加者の条件(出産)' in formatted_data.columns:
        # 列を文字列型に変換してから値を設定（警告を回避）
        formatted_data['参加者の条件(出産)'] = formatted_data['参加者の条件(出産)'].astype(str)
        formatted_data['参加者の条件(出産)'] = '0'
    
    # ステータスの修正（優先順位に従って処理）
    for idx, row in formatted_data.iterrows():
        # 修正・削除新規列の値を正規化
        status_value = normalize_value(row.get('修正・削除新規', ''))
        
        # HP掲載可列の値を正規化
        hp_publish = normalize_value(row.get('HP掲載可', ''))
        
        # 優先順位に従ってステータスを設定
        # 1. 修正・削除新規列の値が「削除」である：ステータス列の値を「private」にする
        if status_value == '削除':
            formatted_data.at[idx, 'ステータス'] = 'private'
        # 2. 修正・削除新規列の値が「削除」でない かつ 空欄である：ステータス列の値を「publish」にする
        elif status_value != '削除' and status_value == '':
            formatted_data.at[idx, 'ステータス'] = 'publish'
        # 3. HP掲載可列の値が1である：ステータス列の値を「publish」にする
        elif hp_publish == '1':
            formatted_data.at[idx, 'ステータス'] = 'publish'
        # 4. HP掲載可列の値が0である：ステータス列の値を「private」にする
        elif hp_publish == '0':
            formatted_data.at[idx, 'ステータス'] = 'private'
        # デフォルト値
        else:
            formatted_data.at[idx, 'ステータス'] = 'publish'
    
    # 順番の修正（実際に変更が必要な行のみ処理）
    formatted_data = formatted_data.reset_index(drop=True)
    
    # 元データとの順番比較用にスラッグをキーとした辞書を作成
    original_order_dict = {}
    for idx, row in original_data.iterrows():
        slug = normalize_value(row.get('スラッグ', ''))
        
        if slug:
            order = normalize_value(row.get('順番', ''))
            original_order_dict[slug] = order
    
    # 新しい順番を設定
    formatted_data['順番'] = range(1, len(formatted_data) + 1)
    
    # 順番の差分チェックと修正・削除新規列の更新（実際に変更があった行のみ）
    for idx, row in formatted_data.iterrows():
        # スラッグの値を正規化
        slug = normalize_value(row.get('スラッグ', ''))
        
        # 現在のステータスを正規化
        current_status = normalize_value(row.get('修正・削除新規', ''))
        
        # すでに「修正」「削除」「新規追加」が入力されている場合は上書きしない
        if current_status in ['修正', '削除', '新規追加']:
            continue
        
        # スラッグが存在し、元データにも存在する場合のみ順番比較
        if slug and slug in original_order_dict:
            # 現在の順番を正規化
            current_order = str(idx + 1)  # 新しい順番（1から始まる連番）
            original_order = original_order_dict[slug]
            
            # 順番が実際に変更された場合のみ「掲載順」を設定
            if current_order != original_order:
                formatted_data.at[idx, '修正・削除新規'] = '掲載順'
    
    return formatted_data

def is_only_account_related_change(main_row, original_data):
    """アカウント関連のみの変更かどうかを判定する関数
    
    Args:
        main_row: メインデータの行
        original_data: 差分検出用データ
    
    Returns:
        bool: アカウント関連のみの変更の場合True
    """
    # スラッグの取得
    slug = str(main_row.get('スラッグ', '')).strip()
    
    if not slug:
        return False
    
    # 差分検出用データから同じスラッグの行を取得
    original_row = original_data[original_data['スラッグ'] == slug]
    
    if original_row.empty:
        return False
    
    original_row = original_row.iloc[0]
    
    # アカウント関連列
    account_columns = ['ｱｶｳﾝﾄ発行有無', 'ｱｶｳﾝﾄ発行年月', 'アカウント発行の登録用メールアドレス']
    
    # アカウント関連以外の列で差分をチェック
    excluded_columns = ['修正・削除新規'] + account_columns
    check_columns = [col for col in main_row.index if col not in excluded_columns]
    
    # アカウント関連以外に変更があるかチェック
    has_non_account_change = False
    for col in check_columns:
        if col in original_row.index:
            main_value = normalize_value(main_row.get(col, ''))
            original_value = normalize_value(original_row.get(col, ''))
            
            if main_value != original_value:
                has_non_account_change = True
                break
    
    # アカウント関連に変更があるかチェック
    has_account_change = False
    for col in account_columns:
        if col in main_row.index and col in original_row.index:
            main_value = normalize_value(main_row.get(col, ''))
            original_value = normalize_value(original_row.get(col, ''))
            
            if main_value != original_value:
                has_account_change = True
                break
    
    # アカウント関連のみの変更の場合：アカウント関連に変更があり、かつアカウント関連以外に変更がない
    return has_account_change and not has_non_account_change

def detect_modified_columns(main_row, original_data, header_mapping):
    """修正対象列を検出する関数（検証関数と同じロジックを使用）
    
    Args:
        main_row: メインデータの行（整形前のデータ）
        original_data: 差分検出用データ
        header_mapping: ヘッダーマッピング辞書
    
    Returns:
        str: 修正された列名のカンマ区切り文字列
    """
    # スラッグの取得（検証関数と同じ処理）
    slug = str(main_row.get('スラッグ', '')).strip()
    
    if not slug:
        return ''
    
    # 差分検出用データから同じスラッグの行を取得
    original_row = original_data[original_data['スラッグ'] == slug]
    
    if original_row.empty:
        return ''
    
    original_row = original_row.iloc[0]
    modified_columns = []
    
    # 修正・削除新規列とアカウント関連列以外の列で差分をチェック（検証関数と同じロジック）
    excluded_columns = ['修正・削除新規', 'ｱｶｳﾝﾄ発行有無', 'ｱｶｳﾝﾄ発行年月', 'アカウント発行の登録用メールアドレス']
    check_columns = [col for col in main_row.index if col not in excluded_columns]
    
    for col in check_columns:
        if col in original_row.index:
            # normalize_value関数を使用して値を正規化（検証関数と同じ処理）
            main_value = normalize_value(main_row.get(col, ''))
            original_value = normalize_value(original_row.get(col, ''))
            
            if main_value != original_value:
                # ヘッダーマッピングがある場合は変換後の名前を使用
                display_col = header_mapping.get(col, col)
                modified_columns.append(display_col)
    
    return ', '.join(modified_columns)

def create_import_files(formatted_data, original_data, user_data, municipality, main_data=None):
    """インポートファイルの作成
    
    Args:
        formatted_data (pd.DataFrame): 整形済みデータ
        original_data (pd.DataFrame): 差分検出用データ
        user_data (pd.DataFrame): ユーザーデータ
        municipality (str): 自治体名
        main_data (pd.DataFrame, optional): 整形前のメインデータ（修正対象列検出用）
    
    Returns:
        dict: 作成されたファイルの辞書
    """
    current_date = datetime.datetime.now().strftime("%Y%m%d")
    current_month = datetime.datetime.now().month
    files = {}
    
    # 育児サークル用データのテンプレートヘッダー
    circle_template_headers = [
        'サークル名', 'スラッグ', 'ステータス', '活動種別', '概要',
        '対象年齢(妊娠前半)', '対象年齢(妊娠後半)', '対象年齢(出産)',
        '対象年齢(0歳)', '対象年齢(1歳前半)', '対象年齢(1歳後半)',
        '対象年齢(2歳前半)', '対象年齢(2歳後半)', '対象年齢(3歳)',
        '対象年齢(4歳)', '対象年齢(5歳)', '対象年齢(6歳(就学前))',
        '対象年齢(6歳(就学後))', '活動場所', '申込方法', '申込方法備考',
        '申込先電話番号', 'Eメールアドレス', '要会費', '会費', 'Webサイト',
        '活動日_営業曜日', '活動日_開始時間', '活動日_終了時間',
        '活動日_営業時間ラベル', '活動日_営業曜日ラベル', '活動日_備考',
        '代表者', '団体名', '団体名（ふりがな）', '幼稚園・保育園チェック',
        '冊子掲載可', 'HP掲載可', 'オープンデータ掲載可', '小学校区',
        '小学校区（ふりがな）', '代表者名', '代表者名（ふりがな）',
        '代表者郵便番号', '代表者住所', '代表者電話番号', '代表者FAX',
        '代表者携帯番号', '記入者', '順番'
    ]
    
    # CSVヘッダーとテンプレートヘッダーのマッピング
    header_mapping = {
        # 参加者の条件系の列名変換
        '参加者の条件(妊娠)': '対象年齢(妊娠前半)',
        '参加者の条件(妊娠後半)': '対象年齢(妊娠後半)',
        '参加者の条件(出産)': '対象年齢(出産)',
        '参加者の条件(0歳)': '対象年齢(0歳)',
        '参加者の条件(1歳)': '対象年齢(1歳前半)',
        '参加者の条件(1歳後半)': '対象年齢(1歳後半)',
        '参加者の条件(2歳)': '対象年齢(2歳前半)',
        '参加者の条件(2歳後半)': '対象年齢(2歳後半)',
        '参加者の条件(3歳)': '対象年齢(3歳)',
        '参加者の条件(4歳)': '対象年齢(4歳)',
        '参加者の条件(5歳)': '対象年齢(5歳)',
        '参加者の条件(6歳)': '対象年齢(6歳(就学前))',
        '参加者の条件(どなたでも)': '対象年齢(6歳(就学後))',
        # その他のヘッダーは同じ名前なのでマッピング不要
    }
    
    # 新規追加の育児サークル
    new_circles = formatted_data[formatted_data['修正・削除新規'] == '新規追加']
    if not new_circles.empty:
        # ヘッダーマッピング（CSVのヘッダーをテンプレートヘッダーに変換）
        # 事前にDataFrameの構造を定義（全て文字列型として初期化）
        new_circles_mapped = pd.DataFrame(index=new_circles.index, 
                                        columns=circle_template_headers, 
                                        dtype=str)
        new_circles_mapped = new_circles_mapped.fillna('')
        
        for template_header in circle_template_headers:
            # マッピングがある場合は元のヘッダー名を使用
            csv_header = None
            for csv_col, template_col in header_mapping.items():
                if template_col == template_header:
                    csv_header = csv_col
                    break
            
            # マッピングがない場合は同じ名前を使用
            if csv_header is None:
                csv_header = template_header
            
            if csv_header in new_circles.columns:
                # 欠損値を適切に処理して代入
                series = new_circles[csv_header].fillna('').astype(str)
                # 'nan'文字列を空文字に置換
                series = series.replace(['nan', 'None', '<NA>'], '')
                new_circles_mapped[template_header] = series
            else:
                new_circles_mapped[template_header] = ''
        
        files[f"{municipality}育児サークル{current_month}月_新規_{current_date}.csv"] = new_circles_mapped
    
    # 修正の育児サークル（明示的に指定された行のみ）
    # 明示的に修正・削除・掲載順が指定されている行のみを修正CSVに含める
    # 暗黙的な修正検出は行わない（インポートデータ整形処理による変更を除外するため）
    # ただし、「修正」の場合はアカウント関連のみの変更は除外する
    candidate_circles = formatted_data[formatted_data['修正・削除新規'].isin(['修正', '削除', '掲載順'])]
    
    # アカウント関連のみの変更を除外
    modified_circles_list = []
    for idx, row in candidate_circles.iterrows():
        status = normalize_value(row.get('修正・削除新規', ''))
        
        if status == '修正':
            # main_dataが提供されている場合、整形前のデータを使用してチェック
            if main_data is not None and idx in main_data.index:
                main_row = main_data.loc[idx]
                # アカウント関連のみの変更の場合は除外
                if is_only_account_related_change(main_row, original_data):
                    continue
            else:
                # main_dataが提供されていない場合、formatted_dataを使用してチェック
                if is_only_account_related_change(row, original_data):
                    continue
        
        # 削除・掲載順の場合、またはアカウント関連以外の変更がある修正の場合は含める
        modified_circles_list.append(idx)
    
    # インデックスリストから該当行を抽出
    if modified_circles_list:
        modified_circles = formatted_data.loc[modified_circles_list]
    else:
        modified_circles = pd.DataFrame()
    if not modified_circles.empty:
        # ヘッダーマッピング（CSVのヘッダーをテンプレートヘッダーに変換）
        # 事前にDataFrameの構造を定義（全て文字列型として初期化）
        modified_circles_mapped = pd.DataFrame(index=modified_circles.index, 
                                             columns=circle_template_headers + ['修正対象列'], 
                                             dtype=str)
        modified_circles_mapped = modified_circles_mapped.fillna('')
        
        for template_header in circle_template_headers:
            # マッピングがある場合は元のヘッダー名を使用
            csv_header = None
            for csv_col, template_col in header_mapping.items():
                if template_col == template_header:
                    csv_header = csv_col
                    break
            
            # マッピングがない場合は同じ名前を使用
            if csv_header is None:
                csv_header = template_header
            
            if csv_header in modified_circles.columns:
                # 欠損値を適切に処理して代入
                series = modified_circles[csv_header].fillna('').astype(str)
                # 'nan'文字列を空文字に置換
                series = series.replace(['nan', 'None', '<NA>'], '')
                modified_circles_mapped[template_header] = series
            else:
                modified_circles_mapped[template_header] = ''
        
        # 修正対象列を検出して追加（整形前のデータを使用）
        for idx, row in modified_circles.iterrows():
            if main_data is not None and idx in main_data.index:
                # 整形前のデータ（main_data）を使用して差分を検出
                main_row = main_data.loc[idx]
                modified_columns = detect_modified_columns(main_row, original_data, header_mapping)
            else:
                # main_dataが提供されていない場合は空文字列
                modified_columns = ''
            modified_circles_mapped.at[idx, '修正対象列'] = modified_columns
        
        # ダウンロード用のデータ（修正対象列を除外）
        download_data = modified_circles_mapped.drop(columns=['修正対象列'])
        
        # ファイル辞書には表示用（修正対象列含む）とダウンロード用（修正対象列除外）の両方を保存
        files[f"{municipality}育児サークル{current_month}月_修正_{current_date}.csv"] = {
            'display_data': modified_circles_mapped,  # 表示用（修正対象列含む）
            'download_data': download_data  # ダウンロード用（修正対象列除外）
        }
    
    # ユーザー新規追加・修正の処理
    user_import_data = create_user_import_data(formatted_data, original_data, user_data)
    if not user_import_data.empty:
        files[f"{municipality}{current_month}月_ユーザー登録{current_date}.csv"] = user_import_data
    
    return files

def create_user_import_data(formatted_data, original_data, user_data):
    """ユーザーインポートデータの作成
    
    Args:
        formatted_data (pd.DataFrame): 整形済みデータ
        original_data (pd.DataFrame): 差分検出用データ
        user_data (pd.DataFrame): ユーザーデータ
    
    Returns:
        pd.DataFrame: ユーザーインポートデータ
    """
    user_import_df = pd.DataFrame(columns=['名前', 'スラッグ', 'メールアドレス', '自己紹介', '種類', 'Webサイト', '画像'])
    
    # アカウント発行有無の条件を正規化して評価
    def is_account_issued(value):
        if pd.isna(value):
            return False
        value_str = str(value).strip()
        if value_str in ['nan', 'None', '<NA>']:
            return False
        return value_str == '○'
    
    # ｱｶｳﾝﾄ発行有無列の差分チェック関数
    def has_account_status_changed(row, original_data):
        """ｱｶｳﾝﾄ発行有無列の値が差分検出用データと異なるかチェック
        
        Returns:
            bool: 以下のパターンで返り値が決まる
                - True: スラッグが存在し、かつｱｶｳﾝﾄ発行有無の値が差分検出用データと異なる場合
                  例: メインデータ「○」→差分検出用データ「空欄」（新規でアカウント発行）
                  例: メインデータ「空欄」→差分検出用データ「○」（アカウント発行取り消し）
                - False: 以下のいずれかの場合
                  1. スラッグが空欄または存在しない場合
                  2. 差分検出用データに該当スラッグが見つからない場合
                  3. ｱｶｳﾝﾄ発行有無の値が差分検出用データと同じ場合
        """
        main_slug = normalize_value(row.get('スラッグ', ''))
        
        if main_slug:  # スラッグが存在する場合のみ処理
            # 差分検出用データから同じスラッグの行を取得
            original_row = original_data[original_data['スラッグ'] == main_slug]
            
            if not original_row.empty:
                # ｱｶｳﾝﾄ発行有無の比較
                main_account_status = is_account_issued(row.get('ｱｶｳﾝﾄ発行有無', ''))
                original_account_status = is_account_issued(original_row.iloc[0].get('ｱｶｳﾝﾄ発行有無', ''))
                
                return main_account_status != original_account_status
        
        return False
    
    # 新規追加のユーザーデータ作成
    # 条件を修正：
    # 条件1（必須）: ｱｶｳﾝﾄ発行有無列 = '○' かつ アカウント発行の登録用メールアドレス列にメールアドレスが記載されている
    # 条件2: 修正・削除新規列の値が「新規追加」である
    # 条件3: ｱｶｳﾝﾄ発行有無列の値が差分検出用データと異なる
    # 
    # 作成されるパターン：
    # - 条件1 かつ 条件2
    # - 条件1 かつ 条件3
    
    # 条件1（必須）: ｱｶｳﾝﾄ発行有無列 = '○' かつ アカウント発行の登録用メールアドレス列にメールアドレスが記載されている
    condition1 = (
        formatted_data['ｱｶｳﾝﾄ発行有無'].apply(is_account_issued) &
        formatted_data['アカウント発行の登録用メールアドレス'].apply(lambda x: normalize_value(x) != '')
    )
    
    # 条件2: 修正・削除新規列の値が「新規追加」である
    condition2 = formatted_data['修正・削除新規'].apply(lambda x: normalize_value(x) == '新規追加')
    
    # 条件3: ｱｶｳﾝﾄ発行有無列の値が差分検出用データと異なる
    condition3 = formatted_data.apply(lambda row: has_account_status_changed(row, original_data), axis=1)
    
    # 条件1が必須で、かつ（条件2または条件3）を満たす行を抽出
    new_accounts = formatted_data[condition1 & (condition2 | condition3)]
    
    # デバッグ情報をセッション状態に保存（表示は後で行う）
    if st.session_state.get('debug_mode', False):
        # アカウント発行有無の状況
        account_issued_count = formatted_data['ｱｶｳﾝﾄ発行有無'].apply(is_account_issued).sum()
        
        # メールアドレス記載の状況
        email_filled_count = formatted_data['アカウント発行の登録用メールアドレス'].apply(lambda x: normalize_value(x) != '').sum()
        
        # 新規追加ステータスの状況
        new_status_count = condition2.sum()
        
        # ｱｶｳﾝﾄ発行有無差分の状況
        account_diff_count = condition3.sum()
        
        # デバッグ情報をセッション状態に保存
        debug_info = {
            'total_rows': len(formatted_data),
            'account_issued_count': account_issued_count,
            'email_filled_count': email_filled_count,
            'new_status_count': new_status_count,
            'account_diff_count': account_diff_count,
            'condition1_count': condition1.sum(),
            'condition2_count': condition2.sum(),
            'condition3_count': condition3.sum(),
            'new_accounts_count': len(new_accounts),
            'new_accounts_sample': new_accounts[['サークル名', 'ｱｶｳﾝﾄ発行有無', 'アカウント発行の登録用メールアドレス', '修正・削除新規']].head() if len(new_accounts) > 0 else None,
            'account_values': formatted_data['ｱｶｳﾝﾄ発行有無'].value_counts() if len(new_accounts) == 0 else None,
        }
        st.session_state.user_csv_debug_info = debug_info
    
    # 修正のユーザーデータ作成（先に実行）
    modified_users_df, modified_row_indices = create_modified_user_data(formatted_data, original_data, user_data)
    
    # 修正対象となった行を新規追加から除外
    if modified_row_indices:
        # 修正対象の行を除外したnew_accountsを作成
        filtered_new_accounts = new_accounts[~new_accounts.index.isin(modified_row_indices)]
    else:
        # 修正対象がない場合は元のnew_accountsをそのまま使用
        filtered_new_accounts = new_accounts
    
    # 新規追加ユーザー情報の詳細を収集（表示用）
    new_user_details = []
    
    # ユーザー作成時のエラー情報を収集
    user_creation_errors = []
    
    # すでに発行済みユーザー情報を収集（エラーではなく情報表示用）
    already_issued_users = []
    
    # 既存のメールアドレスのセットを作成（高速化のため）
    existing_emails = set(user_data['メールアドレス'].astype(str))
    
    # 同じバッチ内でのメールアドレス重複チェック用のセット
    batch_emails = set()
    
    # 新規追加のユーザーデータ作成（修正対象を除外後）
    if not filtered_new_accounts.empty:
        # 既存のスラッグから次の番号を取得
        existing_slugs = user_data['スラッグ'].astype(str)
        cs_numbers = []
        for slug in existing_slugs:
            if slug.startswith('cs') and slug[2:].isdigit():
                num = int(slug[2:])
                if 1 <= num <= 9998:  # cs9999は除外
                    cs_numbers.append(num)
        
        next_number = max(cs_numbers) + 1 if cs_numbers else 1
        
        for idx, row in filtered_new_accounts.iterrows():
            # サークル名の正規化
            raw_circle_name = row.get('サークル名', '')
            if pd.isna(raw_circle_name):
                circle_name = ''
            else:
                circle_name = str(raw_circle_name).strip()
                if circle_name in ['nan', 'None', '<NA>']:
                    circle_name = ''
            
            # メールアドレスの正規化
            raw_email = row.get('アカウント発行の登録用メールアドレス', '')
            if pd.isna(raw_email):
                email = ''
            else:
                email = str(raw_email).strip()
                if email in ['nan', 'None', '<NA>']:
                    email = ''
            
            # ｱｶｳﾝﾄ発行有無の値を取得
            account_issued = is_account_issued(row.get('ｱｶｳﾝﾄ発行有無', ''))
            
            # 修正・削除新規の値を取得
            modification_status = normalize_value(row.get('修正・削除新規', ''))
            
            # 必須項目のチェック
            if not circle_name or not email:
                # 条件1を満たす対象者（ｱｶｳﾝﾄ発行有無=○かつメールアドレス記載予定）に対してのみエラー扱い
                # ただし、既にfiltered_new_accountsで条件1を満たす行のみが抽出されているため、
                # ここに来る行は全て条件1を満たす行である
                missing_fields = []
                if not circle_name:
                    missing_fields.append('サークル名')
                if not email:
                    missing_fields.append('アカウント発行の登録用メールアドレス')
                
                user_creation_errors.append({
                    '行番号': idx + 1,
                    'サークル名': circle_name if circle_name else '（空欄）',
                    'エラー内容': f"{', '.join(missing_fields)}が空欄です",
                    'エラー種別': '必須項目不足'
                })
                continue
            
            # ステータス列による処理分岐
            if modification_status == '新規追加':
                # 新規追加の場合：既存ユーザーとのメールアドレス重複チェック
                if email in existing_emails:
                    user_creation_errors.append({
                        '行番号': idx + 1,
                        'サークル名': circle_name,
                        'エラー内容': f"メールアドレス '{email}' は既に登録されています",
                        'エラー種別': 'メールアドレス重複'
                    })
                    continue
                
                # 同じバッチ内でのメールアドレス重複チェック
                if email in batch_emails:
                    user_creation_errors.append({
                        '行番号': idx + 1,
                        'サークル名': circle_name,
                        'エラー内容': f"メールアドレス '{email}' は同じファイル内の他の行と重複しています",
                        'エラー種別': 'メールアドレス重複'
                    })
                    continue
            else:
                # 新規追加以外の場合：スラッグで紐づいたユーザーとの差分チェック
                slug = normalize_value(row.get('スラッグ', ''))
                if slug:
                    # 代表者スラッグを取得
                    representative_slug = normalize_value(row.get('代表者', ''))
                    if representative_slug:
                        # 既存ユーザーデータから該当ユーザーを検索
                        existing_user = user_data[user_data['スラッグ'] == representative_slug]
                        if not existing_user.empty:
                            existing_user_row = existing_user.iloc[0]
                            existing_user_name = normalize_value(existing_user_row.get('名前', ''))
                            existing_user_email = normalize_value(existing_user_row.get('メールアドレス', ''))
                            
                            # サークル名とメールアドレスの差分チェック
                            name_diff = circle_name != existing_user_name
                            email_diff = email != existing_user_email
                            
                            if not name_diff and not email_diff:
                                # 差分がない場合：「すでに発行済み」として情報表示用に追加
                                already_issued_users.append({
                                    '行番号': idx + 1,
                                    'サークル名': circle_name,
                                    'ユーザースラッグ': representative_slug,
                                    'メールアドレス': existing_user_email,
                                    '詳細': f"名前: {existing_user_name}, メールアドレス: {existing_user_email}"
                                })
                                continue
                            # 差分がある場合はエラーとせず、メインデータを正として処理を継続
            
            # 処理済みメールアドレスとして記録（新規追加の場合のみ）
            if modification_status == '新規追加':
                batch_emails.add(email)
            
            new_slug = f"cs{next_number:04d}"
            
            new_user = {
                '名前': circle_name,
                'スラッグ': new_slug,
                'メールアドレス': email,
                '自己紹介': '',
                '種類': 'blog_writer',
                'Webサイト': '',
                '画像': ''
            }
            
            # 新規追加ユーザー詳細情報を収集
            new_user_details.append({
                'サークル名': circle_name,
                'ユーザースラッグ': new_slug,
                'メールアドレス': email,
                '処理種別': '新規追加',
                '作成理由': f"修正・削除新規: {modification_status}" if modification_status == '新規追加' else 'アカウント発行有無の差分検出'
            })
            
            user_import_df = pd.concat([user_import_df, pd.DataFrame([new_user])], ignore_index=True)
            next_number += 1
    
    # ユーザー作成エラーがある場合は警告情報をセッション状態に保存
    if user_creation_errors:
        error_warning = "### ⚠️ ユーザー作成時にエラーが発生しました\n\n"
        error_warning += f"**{len(user_creation_errors)}件のエラーが見つかりました。以下の行でユーザーが作成されませんでした：**\n\n"
        
        # エラー種別ごとに分類
        missing_fields_errors = [e for e in user_creation_errors if e['エラー種別'] == '必須項目不足']
        duplicate_email_errors = [e for e in user_creation_errors if e['エラー種別'] == 'メールアドレス重複']
        
        if missing_fields_errors:
            error_warning += "**📝 必須項目不足:**\n"
            for error in missing_fields_errors:
                error_warning += f"- 行{error['行番号']}: {error['サークル名']} - {error['エラー内容']}\n"
            error_warning += "\n"
        
        if duplicate_email_errors:
            error_warning += "**📧 メールアドレス重複:**\n"
            for error in duplicate_email_errors:
                error_warning += f"- 行{error['行番号']}: {error['サークル名']} - {error['エラー内容']}\n"
            error_warning += "\n"
        
        error_warning += "**対処方法:**\n"
        error_warning += "1. 必須項目不足：サークル名とアカウント発行の登録用メールアドレスを入力してください\n"
        error_warning += "2. メールアドレス重複：既存と異なるメールアドレスを使用するか、既存ユーザーの修正を検討してください\n"
        error_warning += "3. 同じファイル内での重複：重複するメールアドレスを修正してください\n"
        
        # 警告メッセージをセッション状態に保存
        if 'user_creation_warning' not in st.session_state:
            st.session_state.user_creation_warning = error_warning
    
    # すでに発行済みユーザー情報をセッション状態に保存
    if already_issued_users:
        st.session_state.already_issued_users = already_issued_users
    
    # 新規と修正をマージ
    if not modified_users_df.empty:
        user_import_df = pd.concat([user_import_df, modified_users_df], ignore_index=True)
    
    # 統合ユーザー情報をセッション状態に保存（新規追加と修正の両方）
    all_user_details = []
    
    # 修正情報を追加（既存のuser_modification_detailsから取得）
    if 'user_modification_details' in st.session_state:
        for detail in st.session_state.user_modification_details:
            # 修正情報を統合フォーマットに変換
            changes = []
            if detail['名前変更'] != "変更なし":
                changes.append(f"名前: {detail['名前変更']}")
            if detail['メールアドレス変更'] != "変更なし":
                changes.append(f"メールアドレス: {detail['メールアドレス変更']}")
            
            all_user_details.append({
                'サークル名': detail['サークル名'],
                'ユーザースラッグ': detail['ユーザースラッグ'],
                '処理種別': '修正',
                '変更内容': ', '.join(changes) if changes else '変更なし',
                '処理理由': detail['変更理由']
            })
    
    # 新規追加情報を追加
    for detail in new_user_details:
        all_user_details.append({
            'サークル名': detail['サークル名'],
            'ユーザースラッグ': detail['ユーザースラッグ'],
            '処理種別': detail['処理種別'],
            '変更内容': f"新規作成 (メールアドレス: {detail['メールアドレス']})",
            '処理理由': detail['作成理由']
        })
    
    # 統合情報をセッション状態に保存
    if all_user_details:
        st.session_state.user_comprehensive_details = all_user_details
    
    return user_import_df

def create_modified_user_data(main_data, original_data, user_data):
    """ユーザー修正データの作成
    
    Args:
        main_data (pd.DataFrame): メインデータ
        original_data (pd.DataFrame): 差分検出用データ
        user_data (pd.DataFrame): ユーザーデータ
    
    Returns:
        tuple: (修正されたユーザーデータ, 修正対象行のインデックスリスト)
    """
    modified_users_df = pd.DataFrame(columns=['名前', 'スラッグ', 'メールアドレス', '自己紹介', '種類', 'Webサイト', '画像'])
    
    # 差分表示用のリスト
    modification_details = []
    
    # 修正対象行のインデックスを記録
    modified_row_indices = []
    
    # ユーザー修正時のエラー情報を収集
    user_modification_errors = []
    
    # 1. メインデータのアカウント発行の登録用メールアドレス列の値が差分検出用データと異なる行を抽出
    email_changed_rows = []
    
    for idx, main_row in main_data.iterrows():
        # スラッグでマッチング
        main_slug = normalize_value(main_row.get('スラッグ', ''))
        
        if main_slug:  # スラッグが存在する場合のみ処理
            # 差分検出用データから同じスラッグの行を取得
            original_row = original_data[original_data['スラッグ'] == main_slug]
            
            if not original_row.empty:
                # メールアドレスの比較
                main_email = normalize_value(main_row.get('アカウント発行の登録用メールアドレス', ''))
                original_email = normalize_value(original_row.iloc[0].get('アカウント発行の登録用メールアドレス', ''))
                
                if main_email != original_email:
                    email_changed_rows.append({
                        'index': idx,
                        'main_row': main_row,
                        'main_email': main_email,
                        'original_email': original_email
                    })
    
    # 2. 抽出したデータのうち、「代表者」列の値がユーザーデータの「スラッグ」列と一致するデータを探す
    for change_info in email_changed_rows:
        main_row = change_info['main_row']
        representative_slug = normalize_value(main_row.get('代表者', ''))
        
        if representative_slug:  # 代表者スラッグが存在する場合
            # ユーザーデータから一致するスラッグを探す
            matching_user = user_data[user_data['スラッグ'] == representative_slug]
            
            if not matching_user.empty:
                # 3. 一致するユーザーデータの「名前」「メールアドレス」を更新
                user_row = matching_user.iloc[0]
                
                # メインデータから新しい値を取得
                new_name = normalize_value(main_row.get('サークル名', ''))
                new_email = change_info['main_email']
                
                # 現在のユーザー情報を取得
                current_name = normalize_value(user_row.get('名前', ''))
                current_email = normalize_value(user_row.get('メールアドレス', ''))
                
                # 実際に変更があるかチェック
                name_changed = new_name != current_name
                email_changed = new_email != current_email
                
                if name_changed or email_changed:
                    # メールアドレス重複チェック（自分以外のユーザーとの重複）
                    if email_changed and new_email:
                        # 自分以外のユーザーで同じメールアドレスを持つユーザーがいるかチェック
                        other_users_with_same_email = user_data[
                            (user_data['メールアドレス'] == new_email) & 
                            (user_data['スラッグ'] != representative_slug)
                        ]
                        
                        if not other_users_with_same_email.empty:
                            # メールアドレス重複エラーを記録
                            user_modification_errors.append({
                                '行番号': change_info['index'] + 1,
                                'サークル名': new_name,
                                'ユーザースラッグ': representative_slug,
                                'エラー内容': f"メールアドレス '{new_email}' は他のユーザー（{other_users_with_same_email.iloc[0]['スラッグ']}）が既に使用しています",
                                'エラー種別': 'メールアドレス重複'
                            })
                            continue  # この修正をスキップ
                    
                    # 修正されたユーザーデータを作成
                    modified_user = {
                        '名前': new_name if new_name else current_name,
                        'スラッグ': representative_slug,
                        'メールアドレス': new_email if new_email else current_email,
                        '自己紹介': normalize_value(user_row.get('自己紹介', '')),
                        '種類': normalize_value(user_row.get('種類', '')),
                        'Webサイト': normalize_value(user_row.get('Webサイト', '')),
                        '画像': normalize_value(user_row.get('画像', ''))
                    }
                    
                    modified_users_df = pd.concat([modified_users_df, pd.DataFrame([modified_user])], ignore_index=True)
                    
                    # 修正対象行のインデックスを記録
                    modified_row_indices.append(change_info['index'])
                    
                    # 差分表示用の情報を記録
                    modification_details.append({
                        'サークル名': new_name,
                        'ユーザースラッグ': representative_slug,
                        '名前変更': f"「{current_name}」→「{new_name}」" if name_changed else "変更なし",
                        'メールアドレス変更': f"「{current_email}」→「{new_email}」" if email_changed else "変更なし",
                        '変更理由': 'アカウント発行の登録用メールアドレス列の差分検出'
                    })
    
    # ユーザー修正エラーがある場合は警告情報をセッション状態に保存
    if user_modification_errors:
        error_warning = "### ⚠️ ユーザー修正時にメールアドレス重複エラーが発生しました\n\n"
        error_warning += f"**{len(user_modification_errors)}件のエラーが見つかりました。以下の修正がスキップされました：**\n\n"
        
        for error in user_modification_errors:
            error_warning += f"- 行{error['行番号']}: {error['サークル名']} (ユーザー: {error['ユーザースラッグ']})\n"
            error_warning += f"  - {error['エラー内容']}\n"
        
        error_warning += "\n**対処方法:**\n"
        error_warning += "1. 重複するメールアドレスを修正してください\n"
        error_warning += "2. または、既存ユーザーのメールアドレスを変更してください\n"
        
        # 警告メッセージをセッション状態に保存
        if 'user_modification_warning' not in st.session_state:
            st.session_state.user_modification_warning = error_warning
    
    # 差分を画面表示（Streamlitのセッション状態に保存）
    if modification_details:
        st.session_state.user_modification_details = modification_details
    
    return modified_users_df, modified_row_indices

def validate_facility_csv_file(csv_file):
    """施設情報CSVファイルの検証と読み込みを行う（最適化版）
    
    Args:
        csv_file: アップロードされた施設情報CSVファイル
    
    Returns:
        tuple: (データフレーム, エンコーディング, デバッグ情報)
    
    Raises:
        ValueError: 検証エラーが発生した場合
    """
    import time
    
    # 開始時間を記録
    start_time = time.time()
    debug_info = []
    timing_info = []
    
    # ファイルの内容を一度だけ読み込む（最大サイズを制限）
    file_read_start = time.time()
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    file_content = csv_file.read(MAX_FILE_SIZE)
    if len(file_content) == MAX_FILE_SIZE:
        raise ValueError("ファイルサイズが制限を超えています（最大10MB）")
    
    file_read_time = time.time() - file_read_start
    timing_info.append(f"ファイル読み込み: {file_read_time:.3f}秒")
    
    # chardetによるエンコーディング検出（処理時間短縮のため一時的にコメントアウト）
    # detected_enc, confidence = detect_encoding(file_content)
    # debug_info.append(f"chardetが検出したエンコーディング: {detected_enc} (信頼度: {confidence:.2f})")
    
    # 試行するエンコーディングの順序を決定（固定順序で高速化）
    # encodings = [detected_enc] if detected_enc else []
    # encodings.extend(['utf-8', 'shift-jis', 'cp932', 'euc-jp'])
    # encodings = list(dict.fromkeys(encodings))
    encodings = ['utf-8-sig', 'utf-8', 'shift-jis', 'cp932', 'euc-jp']
    debug_info.append("エンコーディング検出をスキップし、固定順序で試行します（UTF-8 BOM対応）")
    
    encoding_start = time.time()
    successful_encoding = None
    
    for encoding in encodings:
        try:
            encoding_try_start = time.time()
            debug_info.append(f"エンコーディング {encoding} で試行中...")
            
            # まず一部のデータでテスト（先頭1000バイト）
            sample_test_start = time.time()
            sample_size = min(1000, len(file_content))
            try:
                sample = file_content[:sample_size].decode(encoding)
            except UnicodeDecodeError:
                # サンプルテストで失敗した場合、エラー処理付きで再試行
                try:
                    sample = file_content[:sample_size].decode(encoding, errors='ignore')
                    debug_info.append(f"  → サンプルデコードでエラー文字を無視しました")
                except:
                    debug_info.append(f"  → サンプルデコードに失敗")
                    continue
            
            sample_test_time = time.time() - sample_test_start
            timing_info.append(f"サンプルテスト({encoding}): {sample_test_time:.3f}秒")
            
            if not sample.strip():
                debug_info.append(f"  → サンプルデータが空です")
                continue
            
            # 全体をデコード（エラー処理付き）
            full_decode_start = time.time()
            try:
                decoded_content = file_content.decode(encoding)
            except UnicodeDecodeError:
                # エラー文字を無視してデコード
                decoded_content = file_content.decode(encoding, errors='ignore')
                debug_info.append(f"  → 全体デコードでエラー文字を無視しました")
            
            full_decode_time = time.time() - full_decode_start
            timing_info.append(f"全体デコード({encoding}): {full_decode_time:.3f}秒")
            
            # CSVとしての基本検証
            csv_parse_start = time.time()
            df = pd.read_csv(io.StringIO(decoded_content))
            csv_parse_time = time.time() - csv_parse_start
            timing_info.append(f"CSV解析({encoding}): {csv_parse_time:.3f}秒")
            
            # データ品質の検証
            validation_start = time.time()
            if df.empty:
                debug_info.append(f"  → データが空です")
                continue
            if len(df.columns) == 0:
                debug_info.append(f"  → 列が存在しません")
                continue
            
            # CSVファイルとしての基本的な構造確認のみ
            if df.empty:
                debug_info.append(f"  → データが空です")
                continue
            if len(df.columns) == 0:
                debug_info.append(f"  → 列が存在しません")
                continue
            
            validation_time = time.time() - validation_start
            timing_info.append(f"データ検証({encoding}): {validation_time:.3f}秒")
            
            encoding_try_time = time.time() - encoding_try_start
            timing_info.append(f"エンコーディング試行完了({encoding}): {encoding_try_time:.3f}秒")
            
            debug_info.append(f"  → 正常に読み込めました")
            successful_encoding = encoding
            break
            
        except UnicodeDecodeError:
            debug_info.append(f"  → デコードエラー")
            continue
        except pd.errors.EmptyDataError:
            debug_info.append(f"  → 空のCSVファイル")
            raise ValueError("CSVファイルが空です")
        except ValueError as e:
            # 検証エラーは上位に伝播
            raise e
        except Exception as e:
            debug_info.append(f"  → その他のエラー: {str(e)}")
            continue
    
    if successful_encoding is None:
        error_msg = "CSVファイルのエンコーディングを認識できません。以下のいずれかの形式で保存してください：UTF-8、Shift-JIS、CP932、EUC-JP"
        if st.session_state.get('debug_mode', False):
            error_msg += "\n\nデバッグ情報:\n" + "\n".join(debug_info)
        raise ValueError(error_msg)
    
    encoding_total_time = time.time() - encoding_start
    timing_info.append(f"エンコーディング処理合計: {encoding_total_time:.3f}秒")
    
    total_time = time.time() - start_time
    timing_info.append(f"全体処理時間: {total_time:.3f}秒")
    
    # デバッグモード時に処理時間を表示
    if st.session_state.get('debug_mode', False):
        st.write("**⏱️ 施設情報CSVファイル読み込み処理時間:**")
        for timing in timing_info:
            st.write(f"  - {timing}")
    
    return df, successful_encoding, debug_info

def validate_facility_data(df):
    """施設情報データの追加検証を行う
    
    Args:
        df (pd.DataFrame): 検証対象のデータフレーム
    
    Raises:
        ValueError: 検証エラーが発生した場合
    """
    # 施設名の重複チェック
    # NaN値も含めて重複をチェックするため、文字列に変換してから処理
    facility_data_str = df.copy()
    facility_data_str['施設名_str'] = facility_data_str['施設名'].fillna('').astype(str)
    
    # 重複している施設名を検出
    duplicated_mask = facility_data_str['施設名_str'].duplicated(keep=False)
    duplicate_counts = facility_data_str['施設名_str'].value_counts()
    
    # 2回以上出現する施設名を取得
    facility_duplicates = duplicate_counts[duplicate_counts > 1].index.tolist()
    
    if len(facility_duplicates) > 0:
        error_message = "施設情報データ内で重複している施設名が検出されました:\n\n"
        for facility_name_str in facility_duplicates:
            # 重複している施設名のすべての行を取得
            if facility_name_str == '':
                # 空欄・NaN値の場合
                all_duplicate_rows = facility_data_str[facility_data_str['施設名_str'] == '']
                display_name = "（空欄）"
            else:
                all_duplicate_rows = facility_data_str[facility_data_str['施設名_str'] == facility_name_str]
                display_name = facility_name_str
            
            error_message += f"【施設名: {display_name}】\n"
            error_message += f"  重複している行数: {len(all_duplicate_rows)}行\n"
            for idx, row in all_duplicate_rows.iterrows():
                location = row.get('場所', '不明')
                if pd.isna(location):
                    location = '（空欄）'
                # CSVファイルの実際の行番号（ヘッダーを考慮して+2）
                csv_row_number = idx + 2
                error_message += f"  - CSV行{csv_row_number}: 場所=「{location}」\n"
            error_message += "\n"
        
        error_message += "※ 上記の重複行のうち、不要な行を削除してから再度実行してください。"
        raise ValueError(error_message)
    
    # 空欄チェック（施設名が空の行があるかチェック）
    # 重複チェックで既に空欄が検出されている場合はスキップ
    empty_facility_names = df[df['施設名'].isna() | (df['施設名'] == '')]
    if not empty_facility_names.empty and '' not in facility_duplicates:
        error_message = "施設情報データに施設名が空欄の行が存在します:\n\n"
        error_message += f"空欄の行数: {len(empty_facility_names)}行\n"
        for idx, row in empty_facility_names.iterrows():
            location = row.get('場所', '不明')
            if pd.isna(location):
                location = '（空欄）'
            # CSVファイルの実際の行番号（ヘッダーを考慮して+2）
            csv_row_number = idx + 2
            error_message += f"- CSV行{csv_row_number}: 場所=「{location}」\n"
        
        error_message += "\n※ 上記の空欄行を削除または施設名を入力してから再度実行してください。"
        raise ValueError(error_message)

def show_sidebar_usage_guide(selected_function):
    """選択された機能に応じてサイドバーに使い方を表示"""
    st.sidebar.markdown("---")
    
    if selected_function == "データ修正用エクセル作成":
        st.sidebar.markdown("### 📊 データ修正用エクセル作成の使い方")
        st.sidebar.markdown("""
        1. 育児サークルCSVファイルをアップロード
        2. 施設情報CSVファイルをアップロード
        3. 先月分のデータ（Excelファイル）をアップロード
        4. 自治体名を入力（デフォルト：北九州市様）
        5. 「処理開始」ボタンをクリック
        6. 処理が完了したら「処理済みファイルをダウンロード」ボタンが表示される
        7. ダウンロードしたExcelファイルで修正作業を行う
        """)
    else:
        st.sidebar.markdown("### 📋 インポートデータ作成の使い方")
        st.sidebar.markdown("""
        1. 修正済みExcelファイルをアップロード
        2. 必要に応じてスキップする行数を調整
        3. 施設情報CSVファイルとユーザーデータCSVファイルをアップロード
        4. 「データ検証開始」ボタンをクリック
        5. 検証結果を確認
           - エラーがある場合：エラーを修正してから再度検証
           - エラーが0件の場合：次のステップに進む
        6. **エラーが0件の場合のみ**「インポートデータ作成開始」ボタンをクリック
        7. インポートデータが作成されたら、各ファイルをダウンロード
        """)

def show_sidebar_footer():
    """サイドバーの最下段にデバッグモードとバージョン情報を表示"""
    st.sidebar.markdown("---")
    
    # デバッグモードの切り替え
    st.session_state.debug_mode = st.sidebar.checkbox(
        "🔧 デバッグモード", 
        value=st.session_state.debug_mode,
        help="処理時間やセッション状態の詳細情報を表示します"
    )
    
    # デバッグモード時にライブラリバージョン情報を表示
    if st.session_state.debug_mode:
        st.sidebar.markdown("---")
        st.sidebar.markdown("**📚 ライブラリバージョン情報:**")
        
        # 主要ライブラリのバージョンを取得
        try:
            import streamlit as st_lib
            import pandas as pd_lib
            import openpyxl
            import chardet
            import aiohttp
            import requests
            import sys
            
            # バージョン情報を取得
            versions = {
                "Python": sys.version.split()[0],
                "Streamlit": st_lib.__version__,
                "pandas": pd_lib.__version__,
                "openpyxl": openpyxl.__version__,
                "chardet": chardet.__version__,
                "aiohttp": aiohttp.__version__,
                "requests": requests.__version__,
            }
            
            # OpenAIライブラリのバージョンも取得（存在する場合）
            try:
                from openai import __version__ as openai_version
                versions["openai"] = openai_version
            except ImportError:
                versions["openai"] = "未インストール"
            
            # バージョン情報を表示
            for lib, version in versions.items():
                st.sidebar.text(f"  {lib}: {version}")
                
        except Exception as e:
            st.sidebar.error(f"バージョン情報の取得に失敗: {str(e)}")
    
    # バージョン情報（控えめに表示）
    st.sidebar.markdown("---")
    st.sidebar.caption("v2.4 - 2025/07/16")

def setup_page_config():
    """ページの基本設定を行う"""
    st.set_page_config(
        page_title="育児サークル情報処理ツール",
        page_icon="👶",
        layout="wide"
    )

def setup_session_state():
    """セッション状態の初期化"""
    initialize_session_state()
    log_session_state_change("app_started", {})

def main():
    """メイン関数"""
    setup_page_config()
    setup_session_state()
    show_sidebar_chat()
    
    st.title("育児サークル情報処理ツール")
    
    st.header("はじめに", divider='orange')    
    st.write('**←ご利用の手順についてはサイドバーを参照してください。**')
    
    # 機能選択タブ（パフォーマンス向上のため）
    tab1, tab2 = st.tabs(["📊 データ修正用エクセル作成", "📋 インポートデータ作成"])
    
    with tab1:
        # サイドバーの使い方を更新
        show_sidebar_usage_guide("データ修正用エクセル作成")
        show_excel_creation_page()
    
    with tab2:
        # サイドバーの使い方を更新
        show_sidebar_usage_guide("インポートデータ作成")
        show_import_data_page()
    
    # サイドバーの最下段にデバッグモードとバージョン情報を表示
    show_sidebar_footer()

if __name__ == "__main__":
    main() 